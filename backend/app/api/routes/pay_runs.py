"""Pay run lifecycle endpoints.

POST   /payroll/runs                Create a draft pay run
GET    /payroll/runs                List all runs for current user
GET    /payroll/runs/{id}           Get one run
POST   /payroll/runs/preview        Preview a calculation (no DB write)
GET    /payroll/runs/{id}/stubs     List pay stubs for a run
DELETE /payroll/runs/{id}           Delete a draft run

(calculate/finalize/void shipped in the next iteration.)
"""

from decimal import Decimal
from datetime import date, datetime, timezone
from typing import List, Optional, Dict, Any
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, func
from pydantic import BaseModel

from app.db.database import get_db
from app.core.security import get_current_user

from app.models.models import (
    User, PayRun, PayStub, Employee, PayrollSettings,
    YTDBalance, PayrollAuditLog,
)
from app.payroll.service import PayrollService
from app.payroll.types import (
    PayRunEmployeeInput,
    JurisdictionContext,
)


router = APIRouter()


# === Request/Response schemas ===

class CreatePayRunRequest(BaseModel):
    pay_period_start: date
    pay_period_end: date
    pay_date: date
    pay_periods_per_year: int = 26
    country: Optional[str] = None
    subnational: Optional[str] = None
    notes: Optional[str] = None


class PayRunResponse(BaseModel):
    id: str
    pay_period_start: date
    pay_period_end: date
    pay_date: date
    status: str
    country: str
    currency: str
    total_gross: Decimal
    total_deductions: Decimal
    total_net: Decimal
    employee_count: int
    finalized_at: Optional[str] = None
    voided_at: Optional[str] = None
    notes: Optional[str] = None
    created_at: Optional[str] = None


class PreviewRequest(BaseModel):
    pay_period_start: date
    pay_period_end: date
    pay_date: date
    pay_periods_per_year: int = 26
    country: Optional[str] = None
    subnational: Optional[str] = None
    employee_inputs: List[PayRunEmployeeInput]


# === Helpers ===

def _employee_to_dict(emp) -> dict:
    """Convert SQLAlchemy Employee to dict for the service layer."""
    return {
        "id": str(emp.id),
        "first_name": emp.first_name,
        "last_name": emp.last_name,
        "personal_email": emp.personal_email,
        "position_title": emp.position_title,
        "pay_type": emp.pay_type,
        "hourly_rate": str(emp.hourly_rate) if emp.hourly_rate is not None else None,
        "salary_amount": str(emp.salary_amount) if emp.salary_amount is not None else "0",
        "currency": emp.currency or "CAD",
        "tax_info": emp.tax_info or {},
        "vacation_pay_pct": "4.0",
    }


def _run_to_response(run: PayRun) -> PayRunResponse:
    return PayRunResponse(
        id=str(run.id),
        pay_period_start=run.pay_period_start,
        pay_period_end=run.pay_period_end,
        pay_date=run.pay_date,
        status=run.status,
        country=run.country,
        currency=run.currency,
        total_gross=run.total_gross,
        total_deductions=run.total_deductions,
        total_net=run.total_net,
        employee_count=run.employee_count,
        finalized_at=run.finalized_at.isoformat() if run.finalized_at else None,
        voided_at=run.voided_at.isoformat() if run.voided_at else None,
        notes=run.notes,
        created_at=run.created_at.isoformat() if run.created_at else None,
    )


# === Endpoints ===

@router.get("/runs", response_model=List[PayRunResponse])
async def list_pay_runs(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(PayRun)
        .where(PayRun.owner_id == current_user.id)
        .order_by(PayRun.pay_date.desc())
    )
    runs = result.scalars().all()
    return [_run_to_response(r) for r in runs]


@router.post("/runs", response_model=PayRunResponse, status_code=status.HTTP_201_CREATED)
async def create_pay_run(
    body: CreatePayRunRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    settings_result = await db.execute(
        select(PayrollSettings).where(PayrollSettings.owner_id == current_user.id)
    )
    settings = settings_result.scalar_one_or_none()

    country = body.country or (settings.country if settings else "CA")
    currency = settings.currency if settings else "CAD"

    new_run = PayRun(
        owner_id=current_user.id,
        pay_period_start=body.pay_period_start,
        pay_period_end=body.pay_period_end,
        pay_date=body.pay_date,
        status="draft",
        country=country,
        currency=currency,
        total_gross=Decimal("0"),
        total_deductions=Decimal("0"),
        total_net=Decimal("0"),
        employee_count=0,
        notes=body.notes,
    )
    db.add(new_run)
    await db.commit()
    await db.refresh(new_run)
    return _run_to_response(new_run)


@router.get("/runs/{run_id}", response_model=PayRunResponse)
async def get_pay_run(
    run_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(PayRun).where(
            PayRun.id == run_id, PayRun.owner_id == current_user.id
        )
    )
    run = result.scalar_one_or_none()
    if not run:
        raise HTTPException(status_code=404, detail="Pay run not found")
    return _run_to_response(run)


@router.post("/runs/preview")
async def preview_pay_run(
    body: PreviewRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Preview a payroll calculation without persisting anything."""
    settings_result = await db.execute(
        select(PayrollSettings).where(PayrollSettings.owner_id == current_user.id)
    )
    settings = settings_result.scalar_one_or_none()

    country = body.country or (settings.country if settings else None)
    if not country:
        raise HTTPException(
            status_code=400,
            detail="Country required (no PayrollSettings configured)",
        )
    subnational = body.subnational or (settings.province_or_state if settings else None)

    jurisdiction = JurisdictionContext(
        country=country,
        subnational=subnational,
        pay_period_start=body.pay_period_start,
        pay_period_end=body.pay_period_end,
        pay_date=body.pay_date,
        pay_periods_per_year=body.pay_periods_per_year,
    )

    # Load employees referenced in the inputs
    employee_ids = [inp.employee_id for inp in body.employee_inputs]
    emp_result = await db.execute(
        select(Employee).where(
            Employee.id.in_(employee_ids),
            Employee.owner_id == current_user.id,
        )
    )
    employees = emp_result.scalars().all()
    employees_by_id = {str(e.id): _employee_to_dict(e) for e in employees}

    # YTD lookups: empty for now (will be added when finalize lands)
    ytd_by_employee_id = {}

    service = PayrollService()
    try:
        preview = service.preview_run(
            employees_by_id=employees_by_id,
            ytd_by_employee_id=ytd_by_employee_id,
            employee_inputs=body.employee_inputs,
            jurisdiction=jurisdiction,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {
        "total_gross": str(preview.total_gross),
        "total_employee_deductions": str(preview.total_employee_deductions),
        "total_employer_contributions": str(preview.total_employer_contributions),
        "total_net": str(preview.total_net),
        "total_remittance_owed": str(preview.total_remittance_owed),
        "employee_count": preview.employee_count,
        "pay_stubs": [stub.model_dump(mode="json") for stub in preview.pay_stubs],
    }


@router.get("/runs/{run_id}/stubs")
async def list_run_stubs(
    run_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    run_result = await db.execute(
        select(PayRun).where(
            PayRun.id == run_id, PayRun.owner_id == current_user.id
        )
    )
    run = run_result.scalar_one_or_none()
    if not run:
        raise HTTPException(status_code=404, detail="Pay run not found")

    stubs_result = await db.execute(
        select(PayStub).where(PayStub.pay_run_id == run_id)
    )
    stubs = stubs_result.scalars().all()

    return [
        {
            "id": str(s.id),
            "employee_id": str(s.employee_id),
            "employee_name": s.employee_name,
            "gross_pay": str(s.gross_pay),
            "total_employee_deductions": str(s.total_employee_deductions),
            "net_pay": str(s.net_pay),
            "currency": s.currency,
        }
        for s in stubs
    ]


@router.delete("/runs/{run_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_pay_run(
    run_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a draft pay run. Finalized runs cannot be deleted (use void)."""
    result = await db.execute(
        select(PayRun).where(
            PayRun.id == run_id, PayRun.owner_id == current_user.id
        )
    )
    run = result.scalar_one_or_none()
    if not run:
        raise HTTPException(status_code=404, detail="Pay run not found")

    if run.status == "finalized":
        raise HTTPException(
            status_code=400,
            detail="Cannot delete a finalized pay run. Use void instead.",
        )

    await db.delete(run)
    await db.commit()


# ============================================================
# Calculate / Finalize / Void requests
# ============================================================

class CalculateRequest(BaseModel):
    """Body for POST /runs/{id}/calculate.

    employee_inputs: hours and extras per employee in the run.
    pay_periods_per_year: override settings default if provided.
    subnational: override settings.province_or_state if provided.
    """
    employee_inputs: List[PayRunEmployeeInput]
    pay_periods_per_year: Optional[int] = None
    subnational: Optional[str] = None


class VoidRequest(BaseModel):
    reason: str


# ============================================================
# Helpers
# ============================================================

PAY_FREQUENCY_TO_PERIODS = {
    "weekly": 52,
    "bi_weekly": 26,
    "biweekly": 26,
    "semi_monthly": 24,
    "semimonthly": 24,
    "monthly": 12,
}


def _periods_per_year(settings, override: Optional[int]) -> int:
    if override:
        return override
    if settings and settings.default_pay_schedule:
        return PAY_FREQUENCY_TO_PERIODS.get(settings.default_pay_schedule.lower(), 26)
    return 26


def _jurisdiction_key(country: str, subnational: Optional[str]) -> str:
    return f"{country}-{subnational}" if subnational else country


def _ytd_to_dict(ytd: Optional[YTDBalance]) -> Dict[str, Any]:
    if ytd is None:
        return {}
    return {
        "ytd_gross": ytd.ytd_gross,
        "ytd_federal_tax": ytd.ytd_federal_tax,
        "ytd_provincial_or_state_tax": ytd.ytd_provincial_or_state_tax,
        "ytd_social_security_employee": ytd.ytd_social_security_employee,
        "ytd_social_security_employer": ytd.ytd_social_security_employer,
        "ytd_unemployment_employee": ytd.ytd_unemployment_employee,
        "ytd_unemployment_employer": ytd.ytd_unemployment_employer,
        "ytd_pensionable_earnings": ytd.ytd_pensionable_earnings,
        "ytd_insurable_earnings": ytd.ytd_insurable_earnings,
    }


def _apply_ytd_delta(ytd: YTDBalance, stub: PayStub) -> None:
    """Add this pay stub's amounts to the YTD balance."""
    ytd.ytd_gross = (ytd.ytd_gross or Decimal("0")) + stub.gross_pay
    ytd.ytd_federal_tax = (ytd.ytd_federal_tax or Decimal("0")) + stub.federal_tax
    ytd.ytd_provincial_or_state_tax = (
        (ytd.ytd_provincial_or_state_tax or Decimal("0"))
        + stub.provincial_or_state_tax
    )
    ytd.ytd_social_security_employee = (
        (ytd.ytd_social_security_employee or Decimal("0"))
        + stub.social_security_employee
    )
    ytd.ytd_social_security_employer = (
        (ytd.ytd_social_security_employer or Decimal("0"))
        + stub.social_security_employer
    )
    ytd.ytd_unemployment_employee = (
        (ytd.ytd_unemployment_employee or Decimal("0"))
        + stub.unemployment_employee
    )
    ytd.ytd_unemployment_employer = (
        (ytd.ytd_unemployment_employer or Decimal("0"))
        + stub.unemployment_employer
    )
    # Pensionable + insurable from snapshot (engine-dependent)
    snap = stub.calculation_snapshot or {}
    pensionable_after = (
        snap.get("cpp", {}).get("ytd_pensionable_after")
        if isinstance(snap.get("cpp"), dict) else None
    )
    if pensionable_after:
        ytd.ytd_pensionable_earnings = Decimal(str(pensionable_after))

    insurable_after = (
        (snap.get("ei", {}) or {}).get("ytd_insurable_after")
        or (snap.get("fica", {}) or {}).get("new_ytd_medicare")
    )
    if insurable_after:
        ytd.ytd_insurable_earnings = Decimal(str(insurable_after))


def _reverse_ytd_delta(ytd: YTDBalance, stub: PayStub) -> None:
    """Subtract this stub's amounts from YTD (used on void)."""
    ytd.ytd_gross = (ytd.ytd_gross or Decimal("0")) - stub.gross_pay
    ytd.ytd_federal_tax = (ytd.ytd_federal_tax or Decimal("0")) - stub.federal_tax
    ytd.ytd_provincial_or_state_tax = (
        (ytd.ytd_provincial_or_state_tax or Decimal("0"))
        - stub.provincial_or_state_tax
    )
    ytd.ytd_social_security_employee = (
        (ytd.ytd_social_security_employee or Decimal("0"))
        - stub.social_security_employee
    )
    ytd.ytd_social_security_employer = (
        (ytd.ytd_social_security_employer or Decimal("0"))
        - stub.social_security_employer
    )
    ytd.ytd_unemployment_employee = (
        (ytd.ytd_unemployment_employee or Decimal("0"))
        - stub.unemployment_employee
    )
    ytd.ytd_unemployment_employer = (
        (ytd.ytd_unemployment_employer or Decimal("0"))
        - stub.unemployment_employer
    )


# ============================================================
# POST /runs/{id}/calculate
# ============================================================

@router.post("/runs/{run_id}/calculate")
async def calculate_pay_run(
    run_id: UUID,
    body: CalculateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Calculate a pay run and persist pay stubs as draft.

    Idempotent: re-running replaces existing stubs for this run.
    Cannot be called on finalized or voided runs.
    """
    # Load run
    run_result = await db.execute(
        select(PayRun).where(
            PayRun.id == run_id, PayRun.owner_id == current_user.id
        )
    )
    run = run_result.scalar_one_or_none()
    if not run:
        raise HTTPException(404, "Pay run not found")
    if run.status != "draft":
        raise HTTPException(
            400, f"Cannot calculate a run with status '{run.status}' (must be draft)"
        )

    # Load settings
    settings_result = await db.execute(
        select(PayrollSettings).where(PayrollSettings.owner_id == current_user.id)
    )
    settings = settings_result.scalar_one_or_none()

    pay_periods = _periods_per_year(settings, body.pay_periods_per_year)
    subnational = body.subnational or (
        settings.province_or_state if settings else None
    )

    jurisdiction = JurisdictionContext(
        country=run.country,
        subnational=subnational,
        pay_period_start=run.pay_period_start,
        pay_period_end=run.pay_period_end,
        pay_date=run.pay_date,
        pay_periods_per_year=pay_periods,
    )

    # Load employees
    employee_ids = [inp.employee_id for inp in body.employee_inputs]
    emp_result = await db.execute(
        select(Employee).where(
            Employee.id.in_(employee_ids),
            Employee.owner_id == current_user.id,
        )
    )
    employees = emp_result.scalars().all()
    employees_by_id = {str(e.id): _employee_to_dict(e) for e in employees}

    # Load YTD balances (calendar year for v1)
    tax_year = run.pay_period_start.year
    jurisdiction_key = _jurisdiction_key(run.country, subnational)

    ytd_result = await db.execute(
        select(YTDBalance).where(
            YTDBalance.employee_id.in_(employee_ids),
            YTDBalance.tax_year == tax_year,
            YTDBalance.jurisdiction == jurisdiction_key,
        )
    )
    ytd_by_emp = {str(y.employee_id): _ytd_to_dict(y) for y in ytd_result.scalars().all()}

    # Calculate
    service = PayrollService()
    try:
        preview = service.preview_run(
            employees_by_id=employees_by_id,
            ytd_by_employee_id=ytd_by_emp,
            employee_inputs=body.employee_inputs,
            jurisdiction=jurisdiction,
            pay_run_id=str(run.id),
        )
    except ValueError as e:
        raise HTTPException(400, str(e))

    # Preserve any existing memos before wiping so they survive recalculation
    existing_memos_result = await db.execute(
        select(PayStub.employee_id, PayStub.memo)
        .where(PayStub.pay_run_id == run_id, PayStub.memo.isnot(None))
    )
    existing_memos = {str(r.employee_id): r.memo for r in existing_memos_result.all()}

    # Wipe existing pay stubs for this run, then insert fresh
    await db.execute(delete(PayStub).where(PayStub.pay_run_id == run_id))

    for stub in preview.pay_stubs:
        new_stub = PayStub(
            pay_run_id=run.id,
            employee_id=UUID(stub.employee_id),
            employee_name=stub.employee_name,
            employee_email=stub.employee_email,
            position_title=stub.position_title,
            pay_type=stub.pay_type,
            hourly_rate=stub.hourly_rate,
            salary_amount=stub.salary_amount,
            currency=stub.currency,
            hours_regular=stub.hours_regular,
            hours_overtime=stub.hours_overtime,
            hours_stat_holiday=stub.hours_stat_holiday,
            hours_vacation=stub.hours_vacation,
            hours_sick=stub.hours_sick,
            hours_evening=stub.hours_evening,
            hours_overnight=stub.hours_overnight,
            hours_weekend=stub.hours_weekend,
            hours_on_call=stub.hours_on_call,
            hours_travel=stub.hours_travel,
            gross_pay=stub.gross_pay,
            bonus=stub.bonus,
            commission=stub.commission,
            reimbursement=stub.reimbursement,
            federal_tax=stub.federal_tax,
            provincial_or_state_tax=stub.provincial_or_state_tax,
            local_tax=stub.local_tax,
            social_security_employee=stub.social_security_employee,
            social_security_2_employee=stub.social_security_2_employee,
            unemployment_employee=stub.unemployment_employee,
            other_employee_deductions=dict(stub.other_employee_deductions),
            total_employee_deductions=stub.total_employee_deductions,
            social_security_employer=stub.social_security_employer,
            unemployment_employer=stub.unemployment_employer,
            workers_comp_employer=stub.workers_comp_employer,
            other_employer_contributions=dict(stub.other_employer_contributions),
            total_employer_contributions=stub.total_employer_contributions,
            net_pay=stub.net_pay,
            calculation_snapshot=dict(stub.calculation_snapshot),
        )
        # Restore memo if this employee had one from a previous calculation
        preserved_memo = existing_memos.get(str(stub.employee_id))
        if preserved_memo:
            new_stub.memo = preserved_memo
        db.add(new_stub)

    # Update run totals
    run.total_gross = preview.total_gross
    run.total_deductions = preview.total_employee_deductions
    run.total_net = preview.total_net
    run.total_employer_contributions = preview.total_employer_contributions
    run.total_remittance_owed = preview.total_remittance_owed
    run.employee_count = preview.employee_count

    await db.commit()
    await db.refresh(run)

    # Compute change_in_gross_pct: compare each employee's gross to their
    # most recent previous pay stub (any prior pay run of this owner).
    # If no previous stub exists, change is None (Preview shows "New").
    current_employee_ids = [str(s.employee_id) for s in preview.pay_stubs]
    prev_gross_by_emp = {}
    if current_employee_ids:
        prev_result = await db.execute(
            select(PayStub.employee_id, PayStub.gross_pay, PayStub.created_at)
            .where(
                PayStub.employee_id.in_(current_employee_ids),
                PayStub.pay_run_id != run.id,
            )
            .order_by(PayStub.employee_id, PayStub.created_at.desc())
        )
        seen = set()
        for row in prev_result.all():
            emp_id = str(row.employee_id)
            if emp_id in seen:
                continue
            seen.add(emp_id)
            prev_gross_by_emp[emp_id] = Decimal(str(row.gross_pay))

    def _change_pct(current_gross, employee_id):
        prev = prev_gross_by_emp.get(str(employee_id))
        if prev is None or prev == 0:
            return None
        return float((Decimal(str(current_gross)) - prev) / prev)

    # Re-fetch persisted stubs so we have DB ids and memo values
    saved_stubs_result = await db.execute(
        select(PayStub)
        .where(PayStub.pay_run_id == run.id)
        .order_by(PayStub.created_at)
    )
    saved_stubs = saved_stubs_result.scalars().all()

    run_response = _run_to_response(run)
    return {
        **run_response.model_dump(),
        "stubs": [
            {
                "id": str(stub.id),
                "employee_id": str(stub.employee_id),
                "memo": stub.memo,
                "change_in_gross_pct": _change_pct(stub.gross_pay, stub.employee_id),
                "employee_name": stub.employee_name,
                "position_title": stub.position_title,
                "pay_type": stub.pay_type,
                "hourly_rate": str(stub.hourly_rate) if stub.hourly_rate is not None else None,
                "salary_amount": str(stub.salary_amount) if stub.salary_amount is not None else None,
                "hours_regular": str(stub.hours_regular),
                "hours_overtime": str(stub.hours_overtime),
                "hours_stat_holiday": str(stub.hours_stat_holiday),
                "hours_vacation": str(stub.hours_vacation),
                "hours_sick": str(stub.hours_sick),
                "gross_pay": str(stub.gross_pay),
                "federal_tax": str(stub.federal_tax),
                "provincial_or_state_tax": str(stub.provincial_or_state_tax),
                "social_security_employee": str(stub.social_security_employee),
                "social_security_2_employee": str(stub.social_security_2_employee),
                "unemployment_employee": str(stub.unemployment_employee),
                "total_employee_deductions": str(stub.total_employee_deductions),
                "social_security_employer": str(stub.social_security_employer),
                "unemployment_employer": str(stub.unemployment_employer),
                "total_employer_contributions": str(stub.total_employer_contributions),
                "net_pay": str(stub.net_pay),
            }
            for stub in saved_stubs
        ],
    }




# ============================================================
# PATCH /stubs/{stub_id}/memo
# ============================================================

class UpdateStubMemoRequest(BaseModel):
    memo: Optional[str] = None


@router.patch("/stubs/{stub_id}/memo")
async def update_stub_memo(
    stub_id: UUID,
    body: UpdateStubMemoRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update the memo field on a single pay stub.

    Verifies ownership via pay run, then updates memo (max 500 chars).
    """
    result = await db.execute(
        select(PayStub, PayRun)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(
            PayStub.id == stub_id,
            PayRun.owner_id == current_user.id,
        )
    )
    row = result.first()
    if row is None:
        raise HTTPException(status_code=404, detail="Pay stub not found")

    stub = row[0]
    memo = (body.memo or "").strip()
    if len(memo) > 500:
        memo = memo[:500]
    stub.memo = memo if memo else None

    await db.commit()
    await db.refresh(stub)

    return {"id": str(stub.id), "memo": stub.memo}

# ============================================================
# POST /runs/{id}/finalize
# ============================================================

@router.post("/runs/{run_id}/finalize", response_model=PayRunResponse)
async def finalize_pay_run(
    run_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Finalize a pay run: lock it, update YTD balances, write audit log."""
    run_result = await db.execute(
        select(PayRun).where(
            PayRun.id == run_id, PayRun.owner_id == current_user.id
        )
    )
    run = run_result.scalar_one_or_none()
    if not run:
        raise HTTPException(404, "Pay run not found")
    if run.status != "draft":
        raise HTTPException(
            400, f"Only draft runs can be finalized (current: '{run.status}')"
        )

    stubs_result = await db.execute(
        select(PayStub).where(PayStub.pay_run_id == run_id)
    )
    stubs = stubs_result.scalars().all()
    if not stubs:
        raise HTTPException(
            400, "No pay stubs found. Call /calculate first."
        )

    tax_year = run.pay_period_start.year

    for stub in stubs:
        snap = stub.calculation_snapshot or {}
        country = snap.get("country", run.country)
        subnational = snap.get("subnational")
        jurisdiction_key = _jurisdiction_key(country, subnational)

        ytd_result = await db.execute(
            select(YTDBalance).where(
                YTDBalance.employee_id == stub.employee_id,
                YTDBalance.tax_year == tax_year,
                YTDBalance.jurisdiction == jurisdiction_key,
            )
        )
        ytd = ytd_result.scalar_one_or_none()

        if ytd is None:
            ytd = YTDBalance(
                employee_id=stub.employee_id,
                tax_year=tax_year,
                jurisdiction=jurisdiction_key,
            )
            db.add(ytd)

        _apply_ytd_delta(ytd, stub)
        ytd.last_pay_run_id = run.id

    # Lock the run
    run.status = "finalized"
    run.finalized_at = datetime.now(timezone.utc)
    run.finalized_by_user_id = current_user.id

    # Audit
    audit = PayrollAuditLog(
        owner_id=current_user.id,
        entity_type="pay_run",
        entity_id=run.id,
        action="finalized",
        actor_user_id=current_user.id,
        actor_role="admin",
        before_state={"status": "draft"},
        after_state={
            "status": "finalized",
            "employee_count": run.employee_count,
            "total_net": str(run.total_net),
        },
        notes=f"Finalized {len(stubs)} pay stubs",
    )
    db.add(audit)

    await db.commit()
    await db.refresh(run)
    return _run_to_response(run)




# ============================================================
# GET /runs/{id}/done-view - full data for the "Pay run done" page
# ============================================================

@router.get("/runs/{run_id}/done-view")
async def get_pay_run_done_view(
    run_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return everything the Done page needs in one response.

    Includes: run summary + per-employee stubs enriched with employee
    address, YTD balances by tax type, and totals summary.
    """
    # Load the run
    run_result = await db.execute(
        select(PayRun).where(
            PayRun.id == run_id,
            PayRun.owner_id == current_user.id,
        )
    )
    run = run_result.scalar_one_or_none()
    if run is None:
        raise HTTPException(status_code=404, detail="Pay run not found")

    # Load stubs for this run
    stubs_result = await db.execute(
        select(PayStub)
        .where(PayStub.pay_run_id == run_id)
        .order_by(PayStub.employee_name)
    )
    stubs = stubs_result.scalars().all()

    # Load employees (for address, bank info)
    employee_ids = [s.employee_id for s in stubs]
    emp_result = await db.execute(
        select(Employee).where(Employee.id.in_(employee_ids))
    )
    employees_by_id = {str(e.id): e for e in emp_result.scalars().all()}

    # Load YTD balances (calendar year)
    tax_year = run.pay_period_start.year
    ytd_result = await db.execute(
        select(YTDBalance).where(
            YTDBalance.employee_id.in_(employee_ids),
            YTDBalance.tax_year == tax_year,
        )
    )
    ytd_by_employee = {str(y.employee_id): y for y in ytd_result.scalars().all()}

    # Build employee entries
    employees_data = []
    for stub in stubs:
        emp = employees_by_id.get(str(stub.employee_id))
        ytd = ytd_by_employee.get(str(stub.employee_id))

        # Address strings
        address_line1 = emp.address_line1 if emp else None
        address_parts = []
        if emp:
            if emp.city:
                address_parts.append(emp.city)
            if emp.province_or_state:
                address_parts.append(emp.province_or_state)
            if emp.postal_or_zip:
                address_parts.append(emp.postal_or_zip)
        address_line2 = ", ".join(address_parts) if address_parts else None

        # Employee taxes (current + YTD)
        employee_taxes = [
            {
                "type": "Canada Pension Plan",
                "current": str(stub.social_security_employee),
                "ytd": str(ytd.ytd_social_security_employee) if ytd else "0.00",
            },
            {
                "type": "Employment Insurance",
                "current": str(stub.unemployment_employee),
                "ytd": str(ytd.ytd_unemployment_employee) if ytd else "0.00",
            },
            {
                "type": "Income Tax",
                "current": str(stub.federal_tax + stub.provincial_or_state_tax),
                "ytd": str((ytd.ytd_federal_tax + ytd.ytd_provincial_or_state_tax)) if ytd else "0.00",
            },
            {
                "type": "Second Canada Pension Plan",
                "current": str(stub.social_security_2_employee),
                "ytd": "0.00",
            },
        ]

        # Employer contributions (current + YTD)
        employer_contributions = [
            {
                "type": "Employment Insurance Employer",
                "current": str(stub.unemployment_employer),
                "ytd": str(ytd.ytd_unemployment_employer) if ytd else "0.00",
            },
            {
                "type": "Canada Pension Plan Employer",
                "current": str(stub.social_security_employer),
                "ytd": str(ytd.ytd_social_security_employer) if ytd else "0.00",
            },
            {
                "type": "Second Canada Pension Plan Employer",
                "current": "0.00",
                "ytd": "0.00",
            },
        ]

        # Pay lines: derive from hours
        pay_lines = []
        if stub.hours_regular and float(stub.hours_regular) > 0:
            rate = float(stub.hourly_rate) if stub.hourly_rate else 0
            pay_lines.append({
                "type": "Regular Pay",
                "hours": str(stub.hours_regular),
                "rate": str(stub.hourly_rate) if stub.hourly_rate else "0",
                "current": str(round(float(stub.hours_regular) * rate, 2)),
                "ytd": str(ytd.ytd_gross) if ytd else "0.00",
            })
        if stub.hours_overtime and float(stub.hours_overtime) > 0:
            rate = float(stub.hourly_rate) * 1.5 if stub.hourly_rate else 0
            pay_lines.append({
                "type": "Overtime Pay",
                "hours": str(stub.hours_overtime),
                "rate": str(round(rate, 2)),
                "current": str(round(float(stub.hours_overtime) * rate, 2)),
                "ytd": "0.00",
            })
        if stub.hours_stat_holiday and float(stub.hours_stat_holiday) > 0:
            rate = float(stub.hourly_rate) if stub.hourly_rate else 0
            pay_lines.append({
                "type": "Statutory Holiday Pay",
                "hours": str(stub.hours_stat_holiday),
                "rate": str(stub.hourly_rate) if stub.hourly_rate else "0",
                "current": str(round(float(stub.hours_stat_holiday) * rate, 2)),
                "ytd": "0.00",
            })

        employees_data.append({
            "stub_id": str(stub.id),
            "employee_id": str(stub.employee_id),
            "name": stub.employee_name,
            "position_title": stub.position_title,
            "payment_method": "Paper cheque",
            "is_cheque": True,
            "address_line1": address_line1,
            "address_line2": address_line2,
            "paid_from": (emp.bank_name if emp and emp.bank_name else "Employer bank account"),
            "paid_by": f"cheque (${stub.net_pay})",
            "gross_pay": str(stub.gross_pay),
            "employee_deductions": str(stub.total_employee_deductions),
            "employer_cost": str(stub.total_employer_contributions),
            "net_pay": str(stub.net_pay),
            "memo": stub.memo,
            "pay_lines": pay_lines,
            "employee_taxes": employee_taxes,
            "employer_contributions": employer_contributions,
        })

    # Totals
    cheque_count = len(employees_data)  # all cheque for now
    deposit_count = 0
    total_cost = float(run.total_net or 0) + float(run.total_deductions or 0) + float(run.total_employer_contributions or 0)

    return {
        "run": {
            "id": str(run.id),
            "status": run.status,
            "pay_period_start": run.pay_period_start.isoformat() if run.pay_period_start else None,
            "pay_period_end": run.pay_period_end.isoformat() if run.pay_period_end else None,
            "pay_date": run.pay_date.isoformat() if run.pay_date else None,
            "currency": run.currency,
            "total_gross": str(run.total_gross),
            "total_deductions": str(run.total_deductions),
            "total_net": str(run.total_net),
            "total_employer_contributions": str(run.total_employer_contributions or 0),
            "finalized_at": run.finalized_at.isoformat() if run.finalized_at else None,
        },
        "employees": employees_data,
        "totals": {
            "employees_paid": len(employees_data),
            "cheque_count": cheque_count,
            "deposit_count": deposit_count,
            "employee_take_home": str(run.total_net),
            "total_cost": f"{total_cost:.2f}",
            "employee_tax": str(run.total_deductions),
            "employer_tax": str(run.total_employer_contributions or 0),
        },
    }



# ============================================================
# GET /paycheques - list all paycheques (pay stubs) for the user
# ============================================================

@router.get("/paycheques")
async def list_paycheques(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List all paycheques (pay stubs) belonging to the current user.

    Returns each pay stub joined with its pay run for the pay_date and
    run-level status. Only stubs from finalized or voided runs appear;
    stubs from draft runs are hidden (they haven't been issued yet).
    """
    result = await db.execute(
        select(PayStub, PayRun)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(
            PayRun.owner_id == current_user.id,
            PayRun.status.in_(["finalized", "voided"]),
        )
        .order_by(PayRun.pay_date.desc(), PayStub.employee_name)
    )
    rows = result.all()

    paycheques = []
    for stub, run in rows:
        # Map to paycheque status. A stub-level void takes precedence.
        pc_status = "voided" if (stub.voided or run.status == "voided") else "issued"

        paycheques.append({
            "id": str(stub.id),
            "employee_id": str(stub.employee_id),
            "name": stub.employee_name,
            "employee_name": stub.employee_name,
            "pay_date": run.pay_date.isoformat() if run.pay_date else None,
            "pay_period_start": run.pay_period_start.isoformat() if run.pay_period_start else None,
            "pay_period_end": run.pay_period_end.isoformat() if run.pay_period_end else None,
            "total": str(stub.gross_pay),
            "gross_pay": str(stub.gross_pay),
            "net": str(stub.net_pay),
            "net_pay": str(stub.net_pay),
            "pay_method": "cheque",  # defaults per current strategy
            "cheque_number": None,  # not stored yet
            "status": pc_status,
            "currency": run.currency,
            "pay_run_id": str(run.id),
            "memo": stub.memo,
            "is_adjustment": bool(stub.is_adjustment),
            "adjustment_of_stub_id": str(stub.adjustment_of_stub_id) if stub.adjustment_of_stub_id else None,
            "adjustment_reason": stub.adjustment_reason,
        })

    return paycheques



# ============================================================
# GET /paycheques/{stub_id} - full detail for one paycheque
# ============================================================

@router.get("/paycheques/{stub_id}")
async def get_paycheque_detail(
    stub_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return full detail for a single paycheque (pay stub).

    Response shape matches what PaychequeDetail.js expects:
    nested pay / employee_taxes / employer_taxes objects each with
    lines[] and total {current, ytd}.
    """
    result = await db.execute(
        select(PayStub, PayRun)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(
            PayStub.id == stub_id,
            PayRun.owner_id == current_user.id,
        )
    )
    row = result.first()
    if row is None:
        raise HTTPException(status_code=404, detail="Paycheque not found")
    stub, run = row

    # Employee for address + bank
    emp_result = await db.execute(
        select(Employee).where(Employee.id == stub.employee_id)
    )
    emp = emp_result.scalar_one_or_none()

    # YTD balances
    tax_year = run.pay_period_start.year
    ytd_result = await db.execute(
        select(YTDBalance).where(
            YTDBalance.employee_id == stub.employee_id,
            YTDBalance.tax_year == tax_year,
        )
    )
    ytd = ytd_result.scalar_one_or_none()

    # Employer for pay stub header
    from app.models.models import CompanyProfile
    company_res = await db.execute(
        select(CompanyProfile).where(CompanyProfile.user_id == current_user.id)
    )
    company = company_res.scalar_one_or_none()

    # Address
    address_parts = []
    if emp:
        if emp.address_line1:
            address_parts.append(emp.address_line1)
        city_prov_postal = []
        if emp.city: city_prov_postal.append(emp.city)
        if emp.province_or_state: city_prov_postal.append(emp.province_or_state)
        if emp.postal_or_zip: city_prov_postal.append(emp.postal_or_zip)
        if city_prov_postal:
            address_parts.append(", ".join(city_prov_postal))
    address = "\n".join(address_parts) if address_parts else None

    # Pay lines
    pay_lines = []
    if stub.hours_regular and float(stub.hours_regular) > 0:
        rate = float(stub.hourly_rate) if stub.hourly_rate else 0
        pay_lines.append({
            "type": "Regular Pay",
            "hours": str(stub.hours_regular),
            "rate": str(stub.hourly_rate) if stub.hourly_rate else "0",
            "current": str(round(float(stub.hours_regular) * rate, 2)),
            "ytd": str(ytd.ytd_gross) if ytd else "0.00",
        })
    if stub.hours_overtime and float(stub.hours_overtime) > 0:
        rate = float(stub.hourly_rate) * 1.5 if stub.hourly_rate else 0
        pay_lines.append({
            "type": "Overtime Pay",
            "hours": str(stub.hours_overtime),
            "rate": str(round(rate, 2)),
            "current": str(round(float(stub.hours_overtime) * rate, 2)),
            "ytd": "0.00",
        })
    if stub.hours_stat_holiday and float(stub.hours_stat_holiday) > 0:
        rate = float(stub.hourly_rate) if stub.hourly_rate else 0
        pay_lines.append({
            "type": "Stat Holiday Pay",
            "hours": str(stub.hours_stat_holiday),
            "rate": str(stub.hourly_rate) if stub.hourly_rate else "0",
            "current": str(round(float(stub.hours_stat_holiday) * rate, 2)),
            "ytd": "0.00",
        })

    pay_total = {
        "current": str(stub.gross_pay),
        "ytd": str(ytd.ytd_gross) if ytd else "0.00",
    }

    # Employee taxes
    emp_tax_lines = [
        {"type": "Canada Pension Plan", "current": str(stub.social_security_employee), "ytd": str(ytd.ytd_social_security_employee) if ytd else "0.00"},
        {"type": "Employment Insurance", "current": str(stub.unemployment_employee), "ytd": str(ytd.ytd_unemployment_employee) if ytd else "0.00"},
        {"type": "Income Tax", "current": str(stub.federal_tax + stub.provincial_or_state_tax), "ytd": str((ytd.ytd_federal_tax + ytd.ytd_provincial_or_state_tax)) if ytd else "0.00"},
        {"type": "Second Canada Pension Plan", "current": str(stub.social_security_2_employee), "ytd": "0.00"},
    ]
    emp_tax_total = {
        "current": str(stub.total_employee_deductions),
        "ytd": str((ytd.ytd_social_security_employee + ytd.ytd_unemployment_employee + ytd.ytd_federal_tax + ytd.ytd_provincial_or_state_tax)) if ytd else "0.00",
    }

    # Employer taxes
    er_tax_lines = [
        {"type": "Employment Insurance Employer", "current": str(stub.unemployment_employer), "ytd": str(ytd.ytd_unemployment_employer) if ytd else "0.00"},
        {"type": "Canada Pension Plan Employer", "current": str(stub.social_security_employer), "ytd": str(ytd.ytd_social_security_employer) if ytd else "0.00"},
        {"type": "Second Canada Pension Plan Employer", "current": "0.00", "ytd": "0.00"},
    ]
    er_tax_total = {
        "current": str(stub.total_employer_contributions or 0),
        "ytd": str((ytd.ytd_social_security_employer + ytd.ytd_unemployment_employer)) if ytd else "0.00",
    }

    pc_status = "voided" if (stub.voided or run.status == "voided") else "issued"

    return {
        "id": str(stub.id),
        "employee_id": str(stub.employee_id),
        "employee_name": stub.employee_name,
        "employee_address": address,
        "position_title": stub.position_title,
        "pay_date": run.pay_date.isoformat() if run.pay_date else None,
        "pay_period_start": run.pay_period_start.isoformat() if run.pay_period_start else None,
        "pay_period_end": run.pay_period_end.isoformat() if run.pay_period_end else None,
        "paid_from": emp.bank_name if emp and emp.bank_name else "Employer bank account",
        "pay_method": "cheque",
        "cheque_number": None,
        "gross_pay": str(stub.gross_pay),
        "total_pay": str(stub.gross_pay),
        "net_pay": str(stub.net_pay),
        "currency": run.currency,
        "memo": stub.memo,
        "status": pc_status,
        "pay_run_id": str(run.id),
        "employer": {
            "name": company.company_name if company else None,
            "business_number": company.business_number if company else None,
            "payroll_rp_account": company.payroll_rp_account if company else None,
            "address_street": company.address_street if company else None,
            "address_city": company.address_city if company else None,
            "address_province": company.province_state if company else None,
            "address_postal_code": company.address_postal_code if company else None,
        } if company else None,
        "is_adjustment": bool(stub.is_adjustment),
        "adjustment_of_stub_id": str(stub.adjustment_of_stub_id) if stub.adjustment_of_stub_id else None,
        "adjustment_reason": stub.adjustment_reason,
        "pay": {"lines": pay_lines, "total": pay_total},
        "employee_taxes": {"lines": emp_tax_lines, "total": emp_tax_total},
        "employer_taxes": {"lines": er_tax_lines, "total": er_tax_total},
        "deductions_contributions": {"lines": [], "total": None},
    }









# ============================================================
# Payroll Taxes (Payroll Tax Centre) endpoints
# ============================================================

@router.get("/taxes/pd7a")
async def get_pd7a_remittance(
    year: int,
    month: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """PD7A remittance calculation for a single month.

    Aggregates all finalized, non-voided paycheques where pay_date falls
    in the specified month. Returns the totals CRA needs for the PD7A
    Statement of Account.

    Due date is the 15th of the following month (standard monthly remitter
    threshold; quarterly and accelerated remitters have different rules,
    handled in later phases).
    """
    from datetime import date, timedelta
    from calendar import monthrange
    from app.models.models import CompanyProfile

    if year < 2020 or year > 2030:
        raise HTTPException(status_code=400, detail="year must be between 2020 and 2030")
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="month must be between 1 and 12")

    period_start = date(year, month, 1)
    period_end = date(year, month, monthrange(year, month)[1])

    # Aggregate the paycheques
    result = await db.execute(
        select(
            func.count(PayStub.id).label("stub_count"),
            func.count(func.distinct(PayStub.employee_id)).label("employee_count"),
            func.coalesce(func.sum(PayStub.gross_pay), 0).label("gross"),
            func.coalesce(func.sum(PayStub.federal_tax), 0).label("federal_tax"),
            func.coalesce(func.sum(PayStub.provincial_or_state_tax), 0).label("provincial_tax"),
            func.coalesce(func.sum(PayStub.social_security_employee), 0).label("cpp_employee"),
            func.coalesce(func.sum(PayStub.social_security_employer), 0).label("cpp_employer"),
            func.coalesce(func.sum(PayStub.unemployment_employee), 0).label("ei_employee"),
            func.coalesce(func.sum(PayStub.unemployment_employer), 0).label("ei_employer"),
        )
        .select_from(PayStub)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(
            PayRun.owner_id == current_user.id,
            PayRun.pay_date >= period_start,
            PayRun.pay_date <= period_end,
            PayStub.voided == False,
            PayRun.status != "voided",
        )
    )
    row = result.first()

    gross = float(row.gross or 0)
    fed = float(row.federal_tax or 0)
    prov = float(row.provincial_tax or 0)
    cpp_ee = float(row.cpp_employee or 0)
    cpp_er = float(row.cpp_employer or 0)
    ei_ee = float(row.ei_employee or 0)
    ei_er = float(row.ei_employer or 0)

    cpp_total = round(cpp_ee + cpp_er, 2)
    ei_total = round(ei_ee + ei_er, 2)
    tax_total = round(fed + prov, 2)
    current_payment = round(cpp_total + ei_total + tax_total, 2)

    # Due date: 15th of following month
    if month == 12:
        due_date = date(year + 1, 1, 15)
    else:
        due_date = date(year, month + 1, 15)

    today = date.today()
    days_remaining = (due_date - today).days

    if row.stub_count == 0:
        status = "no_activity"
    elif today > due_date:
        status = "overdue"
    else:
        status = "ready_to_pay"

    # Company info
    comp_res = await db.execute(
        select(CompanyProfile).where(CompanyProfile.user_id == current_user.id)
    )
    company = comp_res.scalar_one_or_none()

    account_number_formatted = ""
    if company and company.business_number and company.payroll_rp_account:
        rp = company.payroll_rp_account.replace("RP", "").strip()
        account_number_formatted = f"{company.business_number} RP {rp}"

    return {
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "period_label": f"{period_start.strftime('%d/%m/%Y')} to {period_end.strftime('%d/%m/%Y')}",
        "due_date": due_date.isoformat(),
        "due_date_display": due_date.strftime("%d/%m/%Y"),
        "days_remaining": days_remaining,
        "status": status,
        "paycheque_count": row.stub_count or 0,
        "employee_count": row.employee_count or 0,
        "gross_payroll": round(gross, 2),
        "cpp_employee": round(cpp_ee, 2),
        "cpp_employer": round(cpp_er, 2),
        "cpp_contributions": cpp_total,
        "ei_employee": round(ei_ee, 2),
        "ei_employer": round(ei_er, 2),
        "ei_premiums": ei_total,
        "federal_tax": round(fed, 2),
        "provincial_tax": round(prov, 2),
        "tax_deductions": tax_total,
        "current_payment": current_payment,
        "company": {
            "name": company.company_name if company else "",
            "address_street": company.address_street if company else "",
            "address_city": company.address_city if company else "",
            "address_postal": company.address_postal_code if company else "",
            "business_number": company.business_number if company else "",
            "payroll_rp_account": company.payroll_rp_account if company else "",
            "account_number_formatted": account_number_formatted,
        },
    }



# ============================================================
# Archived tax forms + audit log endpoints
# Country-agnostic; used for PD7A (Canada) today, 941/EPS/etc later
# ============================================================


class ArchiveFormRequest(BaseModel):
    """Body for archiving a form. All the numbers come from the calculator
    endpoint; this just records what the user saw at archive time."""
    country: str = "CA"
    form_type: str  # e.g. "PD7A", "941", "EPS"
    form_subtype: Optional[str] = None  # e.g. "monthly", "quarterly"
    period_start: str  # ISO date
    period_end: str
    period_label: Optional[str] = None
    form_data: dict


@router.post("/taxes/archived-forms")
async def archive_form(
    body: ArchiveFormRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Save a snapshot of a tax form for future viewing/printing.

    Country-agnostic. Body carries country + form_type so this works for
    Canada's PD7A, USA's 941, UK's EPS, etc.
    """
    from datetime import date, datetime
    from uuid import uuid4
    from app.models.models import ArchivedForm
    from app.services.audit_log import log_event

    try:
        period_start = date.fromisoformat(body.period_start)
        period_end = date.fromisoformat(body.period_end)
    except ValueError:
        raise HTTPException(status_code=400, detail="period_start and period_end must be ISO dates")

    if body.country not in ("CA", "US", "UK", "AU", "IE"):
        raise HTTPException(status_code=400, detail="unsupported country: " + body.country)

    archived = ArchivedForm(
        id=uuid4(),
        user_id=current_user.id,
        country=body.country.upper(),
        form_type=body.form_type.upper(),
        form_subtype=body.form_subtype,
        period_start=period_start,
        period_end=period_end,
        period_label=body.period_label,
        form_data=body.form_data,
        archived_by=current_user.id,
    )
    db.add(archived)

    await log_event(
        db,
        user_id=current_user.id,
        event_type="form.archive",
        entity_type="archived_form",
        entity_id=archived.id,
        action="archive",
        details={
            "country": body.country,
            "form_type": body.form_type,
            "period_label": body.period_label,
        },
    )

    await db.commit()
    await db.refresh(archived)

    return {
        "id": str(archived.id),
        "country": archived.country,
        "form_type": archived.form_type,
        "form_subtype": archived.form_subtype,
        "period_start": archived.period_start.isoformat(),
        "period_end": archived.period_end.isoformat(),
        "period_label": archived.period_label,
        "archived_at": archived.archived_at.isoformat(),
    }


@router.get("/taxes/archived-forms")
async def list_archived_forms(
    country: Optional[str] = None,
    form_type: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List archived forms for the current user. Optional filters by country
    and form_type. Sorted newest archived first."""
    from app.models.models import ArchivedForm

    query = select(ArchivedForm).where(ArchivedForm.user_id == current_user.id)
    if country:
        query = query.where(ArchivedForm.country == country.upper())
    if form_type:
        query = query.where(ArchivedForm.form_type == form_type.upper())
    query = query.order_by(ArchivedForm.archived_at.desc())

    result = await db.execute(query)
    rows = result.scalars().all()

    return [
        {
            "id": str(r.id),
            "country": r.country,
            "form_type": r.form_type,
            "form_subtype": r.form_subtype,
            "period_start": r.period_start.isoformat(),
            "period_end": r.period_end.isoformat(),
            "period_label": r.period_label,
            "archived_at": r.archived_at.isoformat(),
        }
        for r in rows
    ]


@router.get("/taxes/archived-forms/{form_id}")
async def get_archived_form(
    form_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get one archived form with full form_data snapshot."""
    from app.models.models import ArchivedForm

    result = await db.execute(
        select(ArchivedForm).where(
            ArchivedForm.id == form_id,
            ArchivedForm.user_id == current_user.id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Archived form not found")

    return {
        "id": str(row.id),
        "country": row.country,
        "form_type": row.form_type,
        "form_subtype": row.form_subtype,
        "period_start": row.period_start.isoformat(),
        "period_end": row.period_end.isoformat(),
        "period_label": row.period_label,
        "form_data": row.form_data,
        "archived_at": row.archived_at.isoformat(),
    }


@router.get("/audit-events")
async def list_audit_events(
    entity_type: Optional[str] = None,
    event_type: Optional[str] = None,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List audit events for the current user. Optional filters.
    Sorted newest first. Max 100 per page (later: cursor pagination)."""
    from app.models.models import AuditEvent

    if limit < 1 or limit > 500:
        limit = 100

    query = select(AuditEvent).where(AuditEvent.user_id == current_user.id)
    if entity_type:
        query = query.where(AuditEvent.entity_type == entity_type)
    if event_type:
        query = query.where(AuditEvent.event_type == event_type)
    query = query.order_by(AuditEvent.created_at.desc()).limit(limit)

    result = await db.execute(query)
    rows = result.scalars().all()

    return [
        {
            "id": str(r.id),
            "event_type": r.event_type,
            "entity_type": r.entity_type,
            "entity_id": str(r.entity_id) if r.entity_id else None,
            "action": r.action,
            "details": r.details or {},
            "created_at": r.created_at.isoformat(),
        }
        for r in rows
    ]

# ============================================================
# GET /paycheques/export/excel - Excel export of paycheques
# ============================================================

@router.get("/paycheques/export/excel")
async def export_paycheques_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export all user paycheques as an Excel (.xlsx) file.

    Columns: Pay Date, Employee, Gross Pay, Net Pay, Status, Type
    Returns a spreadsheet response the browser downloads.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response
    from io import BytesIO
    from datetime import datetime

    result = await db.execute(
        select(PayStub, PayRun)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(PayRun.owner_id == current_user.id)
        .order_by(PayRun.pay_date.desc())
    )
    rows = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Paycheques"

    headers = ["Pay Date", "Employee", "Gross Pay", "Net Pay", "Status", "Type"]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="15A08C", end_color="15A08C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for stub, run in rows:
        pay_date_str = run.pay_date.strftime("%Y-%m-%d") if run.pay_date else ""
        pc_status = "Voided" if (stub.voided or run.status == "voided") else "Issued"
        pc_type = "Adjustment" if stub.is_adjustment else "Regular"
        ws.append([
            pay_date_str,
            stub.employee_name or "",
            float(stub.gross_pay or 0),
            float(stub.net_pay or 0),
            pc_status,
            pc_type,
        ])

    # Set money format on gross/net columns
    for row_i in range(2, ws.max_row + 1):
        ws.cell(row=row_i, column=3).number_format = "$#,##0.00"
        ws.cell(row=row_i, column=4).number_format = "$#,##0.00"

    # Auto-size columns (rough)
    widths = {1: 14, 2: 28, 3: 14, 4: 14, 5: 12, 6: 14}
    for col_i, w in widths.items():
        ws.column_dimensions[get_column_letter(col_i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"paycheques_export_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        }
    )



# ============================================================
# GET /paycheques/export/pdf - PDF export of paycheques list
# ============================================================

@router.get("/paycheques/export/pdf")
async def export_paycheques_pdf(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export all user paycheques as a PDF report.

    Simple table layout: pay date, employee, gross, net, status, type.
    Reuses the same WeasyPrint pipeline as the individual pay stub PDF.
    """
    from weasyprint import HTML
    from fastapi.responses import Response
    from datetime import datetime
    from app.models.models import CompanyProfile

    result = await db.execute(
        select(PayStub, PayRun)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(PayRun.owner_id == current_user.id)
        .order_by(PayRun.pay_date.desc())
    )
    rows = result.all()

    # Company info for header
    company_res = await db.execute(
        select(CompanyProfile).where(CompanyProfile.user_id == current_user.id)
    )
    company = company_res.scalar_one_or_none()
    company_name = company.company_name if company else ""

    def money(v):
        n = float(v or 0)
        return f"${abs(n):,.2f}" if n >= 0 else f"-${abs(n):,.2f}"

    def fmt_date(d):
        return d.strftime("%d %b %Y") if d else ""

    # Build the HTML
    row_html = ""
    for stub, run in rows:
        pc_status = "Voided" if (stub.voided or run.status == "voided") else "Issued"
        pc_type = "Adjustment" if stub.is_adjustment else "Regular"
        row_html += f"""
        <tr>
          <td>{fmt_date(run.pay_date)}</td>
          <td>{stub.employee_name or ""}</td>
          <td class="num">{money(stub.gross_pay)}</td>
          <td class="num">{money(stub.net_pay)}</td>
          <td>{pc_status}</td>
          <td>{pc_type}</td>
        </tr>
        """

    total_gross = sum(float(s.gross_pay or 0) for s, _ in rows)
    total_net = sum(float(s.net_pay or 0) for s, _ in rows)

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
      <meta charset="UTF-8">
      <style>
        @page {{ size: letter; margin: 15mm; }}
        body {{ font-family: Inter, Helvetica, Arial, sans-serif; font-size: 11px; color: #0A0F1A; }}
        .head {{ margin-bottom: 22px; }}
        .company {{ font-size: 14px; font-weight: 700; color: #000000; }}
        .title {{ font-size: 18px; font-weight: 700; color: #000000; margin-top: 4px; }}
        .meta {{ color: #2C3644; font-size: 11px; margin-top: 6px; font-weight: 600; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 6px; }}
        th {{ background: #15A08C; color: white; font-weight: 700; text-align: left; padding: 8px 10px; font-size: 11px; }}
        td {{ padding: 7px 10px; border-bottom: 1px solid #E5E7EB; font-weight: 500; }}
        .num {{ text-align: right; }}
        .totals {{ margin-top: 18px; padding: 12px 14px; background: #E1F5EE; border-radius: 6px; }}
        .totals-row {{ display: flex; justify-content: space-between; font-size: 12px; font-weight: 700; color: #0B7377; }}
      </style>
    </head>
    <body>
      <div class="head">
        <div class="company">{company_name}</div>
        <div class="title">Paycheques Report</div>
        <div class="meta">Generated {datetime.utcnow().strftime('%d %b %Y')}  ·  {len(rows)} paycheque{'s' if len(rows) != 1 else ''}</div>
      </div>
      <table>
        <thead>
          <tr>
            <th>Pay Date</th>
            <th>Employee</th>
            <th class="num">Gross</th>
            <th class="num">Net</th>
            <th>Status</th>
            <th>Type</th>
          </tr>
        </thead>
        <tbody>
          {row_html}
        </tbody>
      </table>
      <div class="totals">
        <div class="totals-row">
          <span>Total Gross</span>
          <span>{money(total_gross)}</span>
        </div>
        <div class="totals-row" style="margin-top: 4px;">
          <span>Total Net</span>
          <span>{money(total_net)}</span>
        </div>
      </div>
    </body>
    </html>
    """

    pdf_bytes = HTML(string=html).write_pdf()
    filename = f"paycheques_report_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
        }
    )

# ============================================================
# GET /paycheques/{stub_id}/pdf - PDF pay stub
# ============================================================

@router.get("/paycheques/{stub_id}/pdf")
async def get_paycheque_pdf(
    stub_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate a PDF pay stub for the given stub.

    Uses WeasyPrint to render an HTML template to PDF. Returns the
    raw PDF as application/pdf. Frontend can open it in a new tab
    or trigger download.
    """
    from weasyprint import HTML
    from jinja2 import Environment, FileSystemLoader
    from fastapi.responses import Response
    from datetime import datetime
    from app.models.models import CompanyProfile
    import os

    # Fetch pay stub + pay run
    result = await db.execute(
        select(PayStub, PayRun)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(
            PayStub.id == stub_id,
            PayRun.owner_id == current_user.id,
        )
    )
    row = result.first()
    if row is None:
        raise HTTPException(status_code=404, detail="Paycheque not found")
    stub, run = row

    # Employee
    emp_result = await db.execute(select(Employee).where(Employee.id == stub.employee_id))
    emp = emp_result.scalar_one_or_none()

    # Employer (company profile)
    company_res = await db.execute(
        select(CompanyProfile).where(CompanyProfile.user_id == current_user.id)
    )
    company = company_res.scalar_one_or_none()

    # YTD
    tax_year = run.pay_period_start.year
    ytd_result = await db.execute(
        select(YTDBalance).where(
            YTDBalance.employee_id == stub.employee_id,
            YTDBalance.tax_year == tax_year,
        )
    )
    ytd = ytd_result.scalar_one_or_none()

    # Money formatter
    def money(v, symbol=False):
        n = float(v or 0)
        body = f"{abs(n):,.2f}"
        sign = "-" if n < 0 else ""
        return f"{sign}${body}" if symbol else f"{sign}{body}"

    def num(v):
        return f"{float(v or 0):,.2f}"

    def fmt_date(d):
        if d is None:
            return ""
        return d.strftime("%d-%m-%Y")

    # Employer lines
    employer_name = company.company_name if company else ""
    employer_line1 = company.address_street if company else ""
    parts = []
    if company:
        if company.address_city:
            parts.append(company.address_city)
        if company.province_state:
            parts.append(company.province_state)
        if company.address_postal_code:
            parts.append(company.address_postal_code)
    employer_line2 = " ".join(parts)

    # Employee address
    address_parts = []
    if emp:
        if emp.address_line1:
            address_parts.append(emp.address_line1)
        cpp = []
        if emp.city: cpp.append(emp.city)
        if emp.province_or_state: cpp.append(emp.province_or_state)
        if emp.postal_or_zip: cpp.append(emp.postal_or_zip)
        if cpp:
            address_parts.append(", ".join(cpp))
    employee_line1 = address_parts[0] if len(address_parts) > 0 else ""
    employee_line2 = address_parts[1] if len(address_parts) > 1 else ""

    # Pay lines
    earnings = []
    total_hours = 0
    if stub.hours_regular and float(stub.hours_regular) > 0:
        rate = float(stub.hourly_rate or 0)
        earnings.append({
            "type": "Regular Pay",
            "hours": num(stub.hours_regular),
            "rate": num(stub.hourly_rate),
            "current": money(float(stub.hours_regular) * rate),
            "ytd": money(ytd.ytd_gross if ytd else 0),
        })
        total_hours += float(stub.hours_regular)
    if stub.hours_overtime and float(stub.hours_overtime) > 0:
        rate = float(stub.hourly_rate or 0) * 1.5
        earnings.append({
            "type": "Overtime Pay",
            "hours": num(stub.hours_overtime),
            "rate": num(rate),
            "current": money(float(stub.hours_overtime) * rate),
            "ytd": money(0),
        })
        total_hours += float(stub.hours_overtime)
    if stub.hours_stat_holiday and float(stub.hours_stat_holiday) > 0:
        rate = float(stub.hourly_rate or 0)
        earnings.append({
            "type": "Stat Holiday Pay",
            "hours": num(stub.hours_stat_holiday),
            "rate": num(stub.hourly_rate),
            "current": money(float(stub.hours_stat_holiday) * rate),
            "ytd": money(0),
        })
        total_hours += float(stub.hours_stat_holiday)

    # Taxes
    taxes = [
        {"type": "Canada Pension Plan", "current": money(stub.social_security_employee), "ytd": money(ytd.ytd_social_security_employee if ytd else 0)},
        {"type": "Employment Insurance", "current": money(stub.unemployment_employee), "ytd": money(ytd.ytd_unemployment_employee if ytd else 0)},
        {"type": "Income Tax", "current": money(stub.federal_tax + stub.provincial_or_state_tax), "ytd": money((ytd.ytd_federal_tax + ytd.ytd_provincial_or_state_tax) if ytd else 0)},
        {"type": "Second Canada Pension Plan", "current": money(stub.social_security_2_employee), "ytd": money(0)},
    ]

    # Summary
    total_pay_current = float(stub.gross_pay or 0)
    total_taxes_current = float(stub.total_employee_deductions or 0)
    deductions_current = 0

    # Load template
    template_dir = os.path.join(os.path.dirname(__file__), "..", "..", "templates")
    env = Environment(loader=FileSystemLoader(template_dir))
    template = env.get_template("pay_stub.html")

    # Adjustment context: fetch original pay date if this is an adjustment
    original_pay_date_display = None
    if stub.is_adjustment and stub.adjustment_of_stub_id:
        orig_res = await db.execute(
            select(PayRun).join(PayStub, PayStub.pay_run_id == PayRun.id).where(PayStub.id == stub.adjustment_of_stub_id)
        )
        orig_run = orig_res.scalar_one_or_none()
        if orig_run and orig_run.pay_date:
            original_pay_date_display = orig_run.pay_date.strftime("%d %b %Y")

    html_content = template.render(
        is_adjustment=bool(stub.is_adjustment),
        adjustment_reason=stub.adjustment_reason,
        original_pay_date=original_pay_date_display,
        employer_name=employer_name,
        employer_line1=employer_line1,
        employer_line2=employer_line2,
        employee_name=stub.employee_name,
        employee_line1=employee_line1,
        employee_line2=employee_line2,
        period_beginning=fmt_date(run.pay_period_start),
        period_ending=fmt_date(run.pay_period_end),
        pay_date=fmt_date(run.pay_date),
        total_hours=num(total_hours),
        net_pay_display=money(stub.net_pay, symbol=True),
        memo=stub.memo,
        earnings=earnings,
        taxes=taxes,
        deductions=[],
        summary_total_pay_current=money(total_pay_current, symbol=True),
        summary_total_pay_ytd=money(ytd.ytd_gross if ytd else 0, symbol=True),
        summary_taxes_current=money(total_taxes_current, symbol=True),
        summary_taxes_ytd=money((ytd.ytd_social_security_employee + ytd.ytd_unemployment_employee + ytd.ytd_federal_tax + ytd.ytd_provincial_or_state_tax) if ytd else 0, symbol=True),
        summary_deductions_current=money(deductions_current, symbol=True),
        summary_deductions_ytd=money(0, symbol=True),
    )

    # Render to PDF
    pdf_bytes = HTML(string=html_content).write_pdf()

    filename = f"paystub_{stub.employee_name.replace(' ', '_')}_{fmt_date(run.pay_date)}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
        }
    )



# ============================================================
# POST /paycheques/{stub_id}/email - Email pay stub with PDF attached
# ============================================================

class EmailPaychequeRequest(BaseModel):
    to_email: str
    subject: Optional[str] = None
    message: Optional[str] = None


@router.post("/paycheques/{stub_id}/email")
async def email_paycheque(
    stub_id: UUID,
    body: EmailPaychequeRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Email the pay stub as a PDF attachment via SendGrid."""
    import sendgrid
    from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
    from app.core.config import settings
    import base64

    # Validate email present
    if not body.to_email or "@" not in body.to_email:
        raise HTTPException(status_code=400, detail="Valid email address required")

    # Fetch paycheque + run + employee to get names for filename/subject
    result = await db.execute(
        select(PayStub, PayRun)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(
            PayStub.id == stub_id,
            PayRun.owner_id == current_user.id,
        )
    )
    row = result.first()
    if row is None:
        raise HTTPException(status_code=404, detail="Paycheque not found")
    stub, run = row

    # Get PDF bytes by calling internal PDF generator inline
    # (Reuse the same rendering logic as the /pdf endpoint)
    from weasyprint import HTML
    from jinja2 import Environment, FileSystemLoader
    from app.models.models import CompanyProfile
    import os

    emp_res = await db.execute(select(Employee).where(Employee.id == stub.employee_id))
    emp = emp_res.scalar_one_or_none()

    company_res = await db.execute(
        select(CompanyProfile).where(CompanyProfile.user_id == current_user.id)
    )
    company = company_res.scalar_one_or_none()

    tax_year = run.pay_period_start.year
    ytd_res = await db.execute(
        select(YTDBalance).where(
            YTDBalance.employee_id == stub.employee_id,
            YTDBalance.tax_year == tax_year,
        )
    )
    ytd = ytd_res.scalar_one_or_none()

    def money(v, symbol=False):
        n = float(v or 0)
        s = f"{abs(n):,.2f}"
        sign = "-" if n < 0 else ""
        return f"{sign}${s}" if symbol else f"{sign}{s}"

    def num(v):
        return f"{float(v or 0):,.2f}"

    def fmt_date(d):
        return d.strftime("%d-%m-%Y") if d else ""

    employer_name = company.company_name if company else ""
    employer_line1 = company.address_street if company else ""
    parts = []
    if company:
        if company.address_city: parts.append(company.address_city)
        if company.province_state: parts.append(company.province_state)
        if company.address_postal_code: parts.append(company.address_postal_code)
    employer_line2 = " ".join(parts)

    address_parts = []
    if emp:
        if emp.address_line1: address_parts.append(emp.address_line1)
        cpp = []
        if emp.city: cpp.append(emp.city)
        if emp.province_or_state: cpp.append(emp.province_or_state)
        if emp.postal_or_zip: cpp.append(emp.postal_or_zip)
        if cpp: address_parts.append(", ".join(cpp))
    employee_line1 = address_parts[0] if address_parts else ""
    employee_line2 = address_parts[1] if len(address_parts) > 1 else ""

    earnings = []
    total_hours = 0
    if stub.hours_regular and float(stub.hours_regular) > 0:
        rate = float(stub.hourly_rate or 0)
        earnings.append({"type":"Regular Pay","hours":num(stub.hours_regular),"rate":num(stub.hourly_rate),"current":money(float(stub.hours_regular) * rate),"ytd":money(ytd.ytd_gross if ytd else 0)})
        total_hours += float(stub.hours_regular)
    if stub.hours_overtime and float(stub.hours_overtime) > 0:
        rate = float(stub.hourly_rate or 0) * 1.5
        earnings.append({"type":"Overtime Pay","hours":num(stub.hours_overtime),"rate":num(rate),"current":money(float(stub.hours_overtime) * rate),"ytd":money(0)})
        total_hours += float(stub.hours_overtime)
    if stub.hours_stat_holiday and float(stub.hours_stat_holiday) > 0:
        rate = float(stub.hourly_rate or 0)
        earnings.append({"type":"Stat Holiday Pay","hours":num(stub.hours_stat_holiday),"rate":num(stub.hourly_rate),"current":money(float(stub.hours_stat_holiday) * rate),"ytd":money(0)})
        total_hours += float(stub.hours_stat_holiday)

    taxes = [
        {"type":"Canada Pension Plan","current":money(stub.social_security_employee),"ytd":money(ytd.ytd_social_security_employee if ytd else 0)},
        {"type":"Employment Insurance","current":money(stub.unemployment_employee),"ytd":money(ytd.ytd_unemployment_employee if ytd else 0)},
        {"type":"Income Tax","current":money(stub.federal_tax + stub.provincial_or_state_tax),"ytd":money((ytd.ytd_federal_tax + ytd.ytd_provincial_or_state_tax) if ytd else 0)},
        {"type":"Second Canada Pension Plan","current":money(stub.social_security_2_employee),"ytd":money(0)},
    ]

    total_pay_current = float(stub.gross_pay or 0)
    total_taxes_current = float(stub.total_employee_deductions or 0)

    template_dir = os.path.join(os.path.dirname(__file__), "..", "..", "templates")
    env = Environment(loader=FileSystemLoader(template_dir))
    template = env.get_template("pay_stub.html")
    html_content = template.render(
        employer_name=employer_name, employer_line1=employer_line1, employer_line2=employer_line2,
        employee_name=stub.employee_name, employee_line1=employee_line1, employee_line2=employee_line2,
        period_beginning=fmt_date(run.pay_period_start), period_ending=fmt_date(run.pay_period_end),
        pay_date=fmt_date(run.pay_date), total_hours=num(total_hours),
        net_pay_display=money(stub.net_pay, symbol=True), memo=stub.memo,
        earnings=earnings, taxes=taxes, deductions=[],
        summary_total_pay_current=money(total_pay_current, symbol=True),
        summary_total_pay_ytd=money(ytd.ytd_gross if ytd else 0, symbol=True),
        summary_taxes_current=money(total_taxes_current, symbol=True),
        summary_taxes_ytd=money((ytd.ytd_social_security_employee + ytd.ytd_unemployment_employee + ytd.ytd_federal_tax + ytd.ytd_provincial_or_state_tax) if ytd else 0, symbol=True),
        summary_deductions_current=money(0, symbol=True), summary_deductions_ytd=money(0, symbol=True),
    )
    pdf_bytes = HTML(string=html_content).write_pdf()

    # Build subject and message defaults if not provided
    period_display = f"{run.pay_period_start.strftime('%d %b %Y')} - {run.pay_period_end.strftime('%d %b %Y')}"
    subject = body.subject or f"Your pay stub for {period_display}"
    message = body.message or f"Hi {stub.employee_name.split()[0]},\n\nYour pay stub for the pay period {period_display} is attached.\n\nIf you have any questions, please reach out.\n\nThanks,\n{employer_name or current_user.email}"

    # Build filename
    safe_name = stub.employee_name.replace(" ", "_")
    filename = f"paystub_{safe_name}_{fmt_date(run.pay_date)}.pdf"

    # SendGrid mail
    mail = Mail(
        from_email=(settings.sendgrid_from_email, settings.sendgrid_from_name),
        to_emails=body.to_email,
        subject=subject,
        plain_text_content=message,
    )
    encoded = base64.b64encode(pdf_bytes).decode()
    attachment = Attachment(
        FileContent(encoded),
        FileName(filename),
        FileType("application/pdf"),
        Disposition("attachment"),
    )
    mail.attachment = attachment

    try:
        sg = sendgrid.SendGridAPIClient(api_key=settings.sendgrid_api_key)
        response = sg.send(mail)
        if response.status_code >= 300:
            raise HTTPException(status_code=502, detail=f"SendGrid returned {response.status_code}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to send email: {str(e)}")

    return {"sent": True, "to": body.to_email}



# ============================================================
# POST /paycheques/{stub_id}/adjust - Create adjustment cheque
# ============================================================

class AdjustPaychequeRequest(BaseModel):
    direction: str  # "extra_pay" or "recover_overpayment"
    gross_amount: float  # always positive
    reason: str


@router.post("/paycheques/{stub_id}/adjust")
async def adjust_paycheque(
    stub_id: UUID,
    body: AdjustPaychequeRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create an adjustment cheque linked to the given original pay stub.

    Business rules:
    - Original stub must be finalized, not voided, not itself an adjustment
    - direction 'extra_pay': positive amount, employee owed more
    - direction 'recover_overpayment': negative amount, employee owes back
    - Taxes calculated on the signed amount via the same tax engine as
      regular pay runs, using the original pay run's dates so YTD/period
      context matches
    - Adjustment saved as new PayStub row with is_adjustment=true
    - YTD balances updated by the delta (positive or negative)
    """
    from decimal import Decimal
    from app.payroll.service import PayrollService
    from app.payroll.types import PayRunEmployeeInput, HoursWorked, JurisdictionContext

    if body.direction not in ("extra_pay", "recover_overpayment"):
        raise HTTPException(status_code=400, detail="direction must be 'extra_pay' or 'recover_overpayment'")
    if not body.reason or not body.reason.strip():
        raise HTTPException(status_code=400, detail="reason is required")
    if body.gross_amount <= 0:
        raise HTTPException(status_code=400, detail="gross_amount must be positive; direction controls sign")

    # Fetch original stub + run
    result = await db.execute(
        select(PayStub, PayRun)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(PayStub.id == stub_id, PayRun.owner_id == current_user.id)
    )
    row = result.first()
    if row is None:
        raise HTTPException(status_code=404, detail="Original paycheque not found")
    orig_stub, orig_run = row

    if orig_run.status != "finalized":
        raise HTTPException(status_code=400, detail="Can only adjust finalized paycheques")
    if orig_stub.voided:
        raise HTTPException(status_code=400, detail="Cannot adjust a voided paycheque")
    if orig_stub.is_adjustment:
        raise HTTPException(status_code=400, detail="Cannot create adjustment on an adjustment cheque")

    # Fetch employee
    emp_result = await db.execute(select(Employee).where(Employee.id == orig_stub.employee_id))
    emp = emp_result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Fetch YTD (should exist for finalized run)
    tax_year = orig_run.pay_period_start.year
    ytd_result = await db.execute(
        select(YTDBalance).where(
            YTDBalance.employee_id == orig_stub.employee_id,
            YTDBalance.tax_year == tax_year,
        )
    )
    ytd = ytd_result.scalar_one_or_none()

    # Sign the amount by direction
    signed_gross = Decimal(str(body.gross_amount))
    if body.direction == "recover_overpayment":
        signed_gross = -signed_gross

    # Pay periods per year map
    freq_map = {"weekly": 52, "biweekly": 26, "semimonthly": 24, "monthly": 12}
    pay_periods_per_year = freq_map.get(emp.pay_frequency or "biweekly", 26)

    # Build employee dict (matches shape the engine expects)
    employee_dict = {
        "id": str(emp.id),
        "first_name": emp.first_name,
        "last_name": emp.last_name,
        "personal_email": emp.personal_email,
        "position_title": emp.position_title,
        "pay_type": emp.pay_type,
        "hourly_rate": float(emp.hourly_rate) if emp.hourly_rate else 0,
        "salary_amount": float(emp.salary_amount) if emp.salary_amount else 0,
        "currency": emp.currency or "CAD",
        "country": emp.country or "CA",
        "province_or_state": emp.province_or_state or "AB",
        "pay_frequency": emp.pay_frequency or "biweekly",
        "tax_info": emp.tax_info or {},
    }

    hours_input = PayRunEmployeeInput(
        employee_id=str(emp.id),
        hours=HoursWorked(),  # all zeros
        bonus=signed_gross,
        commission=Decimal("0"),
        reimbursement=Decimal("0"),
    )

    ytd_dict = {
        "ytd_gross": str(ytd.ytd_gross) if ytd else "0",
        "ytd_social_security_employee": str(ytd.ytd_social_security_employee) if ytd else "0",
        "ytd_unemployment_employee": str(ytd.ytd_unemployment_employee) if ytd else "0",
        "ytd_federal_tax": str(ytd.ytd_federal_tax) if ytd else "0",
        "ytd_provincial_or_state_tax": str(ytd.ytd_provincial_or_state_tax) if ytd else "0",
    }

    # Use two-letter province code, not full name
    prov_code_map = {"Alberta": "AB", "British Columbia": "BC", "Manitoba": "MB",
                     "New Brunswick": "NB", "Newfoundland and Labrador": "NL",
                     "Nova Scotia": "NS", "Ontario": "ON", "Prince Edward Island": "PE",
                     "Quebec": "QC", "Saskatchewan": "SK", "Northwest Territories": "NT",
                     "Nunavut": "NU", "Yukon": "YT"}
    prov = emp.province_or_state or "AB"
    prov = prov_code_map.get(prov, prov)

    jurisdiction = JurisdictionContext(
        country=emp.country or "CA",
        province_or_state=prov,
        pay_frequency=emp.pay_frequency or "biweekly",
        tax_year=tax_year,
        pay_period_start=orig_run.pay_period_start,
        pay_period_end=orig_run.pay_period_end,
        pay_date=orig_run.pay_date,
        pay_periods_per_year=pay_periods_per_year,
    )

    service = PayrollService()
    calc_result = service.calculate_one_employee(
        employee=employee_dict,
        hours_input=hours_input,
        ytd=ytd_dict,
        jurisdiction=jurisdiction,
    )

    # Create adjustment PayStub row
    from uuid import uuid4
    adjustment_stub = PayStub(
        id=uuid4(),
        pay_run_id=orig_stub.pay_run_id,
        employee_id=orig_stub.employee_id,
        employee_name=orig_stub.employee_name,
        employee_email=orig_stub.employee_email,
        position_title=orig_stub.position_title,
        pay_type=orig_stub.pay_type,
        hourly_rate=orig_stub.hourly_rate,
        salary_amount=Decimal("0"),
        currency=orig_stub.currency,
        hours_regular=Decimal("0"),
        hours_overtime=Decimal("0"),
        hours_stat_holiday=Decimal("0"),
        hours_vacation=Decimal("0"),
        hours_sick=Decimal("0"),
        hours_evening=Decimal("0"),
        hours_overnight=Decimal("0"),
        hours_weekend=Decimal("0"),
        hours_on_call=Decimal("0"),
        hours_travel=Decimal("0"),
        gross_pay=calc_result.gross_pay,
        bonus=signed_gross,
        commission=Decimal("0"),
        reimbursement=Decimal("0"),
        federal_tax=calc_result.federal_tax,
        provincial_or_state_tax=calc_result.provincial_or_state_tax,
        social_security_employee=calc_result.social_security_employee,
        social_security_employer=calc_result.social_security_employer,
        social_security_2_employee=calc_result.social_security_2_employee,
        unemployment_employee=calc_result.unemployment_employee,
        unemployment_employer=calc_result.unemployment_employer,
        total_employee_deductions=calc_result.total_employee_deductions,
        total_employer_contributions=calc_result.total_employer_contributions,
        net_pay=calc_result.net_pay,
        is_adjustment=True,
        adjustment_of_stub_id=orig_stub.id,
        adjustment_reason=body.reason.strip(),
        memo=f"Adjustment: {body.reason.strip()[:80]}",
        voided=False,
    )
    db.add(adjustment_stub)

    # Update YTD balances by delta (may create if missing)
    if ytd is None:
        ytd = YTDBalance(
            id=uuid4(),
            employee_id=orig_stub.employee_id,
            tax_year=tax_year,
            jurisdiction=(emp.country or "CA") + "-" + prov,
            ytd_gross=Decimal("0"),
            ytd_federal_tax=Decimal("0"),
            ytd_provincial_or_state_tax=Decimal("0"),
            ytd_social_security_employee=Decimal("0"),
            ytd_unemployment_employee=Decimal("0"),
        )
        db.add(ytd)

    ytd.ytd_gross = (ytd.ytd_gross or Decimal("0")) + calc_result.gross_pay
    ytd.ytd_social_security_employee = (ytd.ytd_social_security_employee or Decimal("0")) + calc_result.social_security_employee
    ytd.ytd_unemployment_employee = (ytd.ytd_unemployment_employee or Decimal("0")) + calc_result.unemployment_employee
    ytd.ytd_federal_tax = (ytd.ytd_federal_tax or Decimal("0")) + calc_result.federal_tax
    ytd.ytd_provincial_or_state_tax = (ytd.ytd_provincial_or_state_tax or Decimal("0")) + calc_result.provincial_or_state_tax

    await db.commit()
    await db.refresh(adjustment_stub)

    return {
        "adjustment_stub_id": str(adjustment_stub.id),
        "direction": body.direction,
        "gross_amount": float(calc_result.gross_pay),
        "net_amount": float(calc_result.net_pay),
        "taxes": {
            "federal": float(calc_result.federal_tax),
            "provincial": float(calc_result.provincial_or_state_tax),
            "cpp": float(calc_result.social_security_employee),
            "ei": float(calc_result.unemployment_employee),
        },
        "reason": body.reason.strip(),
    }

# ============================================================
# POST /paycheques/{stub_id}/void
# ============================================================

class VoidPaychequeRequest(BaseModel):
    reason: str


@router.post("/paycheques/{stub_id}/void")
async def void_paycheque(
    stub_id: UUID,
    body: VoidPaychequeRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Void a single paycheque (pay stub).

    Requirements per V1 spec:
    - Reason required (non-empty after strip)
    - Only stubs from finalized runs can be voided
    - Can't void an already-voided stub
    - Reverses YTD balances for the employee (subtracts stub's amounts)
    - Marks the stub as voided (does NOT touch other stubs in the run)
    - Writes audit log with reason
    """
    reason = (body.reason or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="Reason is required to void a paycheque")

    # Load stub + run to verify ownership
    result = await db.execute(
        select(PayStub, PayRun)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(
            PayStub.id == stub_id,
            PayRun.owner_id == current_user.id,
        )
    )
    row = result.first()
    if row is None:
        raise HTTPException(status_code=404, detail="Paycheque not found")
    stub, run = row

    if stub.voided:
        raise HTTPException(status_code=400, detail="This paycheque is already voided")

    if run.status != "finalized":
        raise HTTPException(
            status_code=400,
            detail=f"Only paycheques from finalized runs can be voided (current: {run.status})"
        )

    # Reverse YTD balances for this employee (this year, this jurisdiction)
    snap = stub.calculation_snapshot or {}
    country = snap.get("country", run.country)
    subnational = snap.get("subnational")
    jurisdiction_key = _jurisdiction_key(country, subnational)
    tax_year = run.pay_period_start.year

    ytd_result = await db.execute(
        select(YTDBalance).where(
            YTDBalance.employee_id == stub.employee_id,
            YTDBalance.tax_year == tax_year,
            YTDBalance.jurisdiction == jurisdiction_key,
        )
    )
    ytd = ytd_result.scalar_one_or_none()
    if ytd is not None:
        # Subtract this stub's amounts from YTD
        ytd.ytd_gross = (ytd.ytd_gross or Decimal(0)) - (stub.gross_pay or Decimal(0))
        ytd.ytd_federal_tax = (ytd.ytd_federal_tax or Decimal(0)) - (stub.federal_tax or Decimal(0))
        ytd.ytd_provincial_or_state_tax = (ytd.ytd_provincial_or_state_tax or Decimal(0)) - (stub.provincial_or_state_tax or Decimal(0))
        ytd.ytd_social_security_employee = (ytd.ytd_social_security_employee or Decimal(0)) - (stub.social_security_employee or Decimal(0))
        ytd.ytd_social_security_employer = (ytd.ytd_social_security_employer or Decimal(0)) - (stub.social_security_employer or Decimal(0))
        ytd.ytd_unemployment_employee = (ytd.ytd_unemployment_employee or Decimal(0)) - (stub.unemployment_employee or Decimal(0))
        ytd.ytd_unemployment_employer = (ytd.ytd_unemployment_employer or Decimal(0)) - (stub.unemployment_employer or Decimal(0))

    # Mark the stub voided
    stub.voided = True
    stub.voided_at = datetime.now(timezone.utc)
    stub.voided_reason = reason[:1000]

    # Audit log
    audit = PayrollAuditLog(
        owner_id=current_user.id,
        entity_type="pay_stub",
        entity_id=stub.id,
        action="voided",
        actor_user_id=current_user.id,
        actor_role="admin",
        before_state={"voided": False},
        after_state={
            "voided": True,
            "reason": reason,
            "gross_pay_reversed": str(stub.gross_pay),
            "net_pay_reversed": str(stub.net_pay),
        },
        notes=f"Paycheque voided: {reason[:200]}",
    )
    db.add(audit)

    await db.commit()
    await db.refresh(stub)

    return {
        "id": str(stub.id),
        "voided": stub.voided,
        "voided_at": stub.voided_at.isoformat() if stub.voided_at else None,
        "voided_reason": stub.voided_reason,
        "status": "voided",
    }

# ============================================================
# POST /runs/{id}/void
# ============================================================

@router.post("/runs/{run_id}/void", response_model=PayRunResponse)
async def void_pay_run(
    run_id: UUID,
    body: VoidRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Void a pay run. If finalized, reverses YTD updates."""
    run_result = await db.execute(
        select(PayRun).where(
            PayRun.id == run_id, PayRun.owner_id == current_user.id
        )
    )
    run = run_result.scalar_one_or_none()
    if not run:
        raise HTTPException(404, "Pay run not found")
    if run.status == "voided":
        raise HTTPException(400, "Pay run is already voided")

    was_finalized = run.status == "finalized"

    if was_finalized:
        # Reverse YTD updates
        stubs_result = await db.execute(
            select(PayStub).where(PayStub.pay_run_id == run_id)
        )
        stubs = stubs_result.scalars().all()

        tax_year = run.pay_period_start.year
        for stub in stubs:
            snap = stub.calculation_snapshot or {}
            country = snap.get("country", run.country)
            subnational = snap.get("subnational")
            jurisdiction_key = _jurisdiction_key(country, subnational)

            ytd_result = await db.execute(
                select(YTDBalance).where(
                    YTDBalance.employee_id == stub.employee_id,
                    YTDBalance.tax_year == tax_year,
                    YTDBalance.jurisdiction == jurisdiction_key,
                )
            )
            ytd = ytd_result.scalar_one_or_none()
            if ytd:
                _reverse_ytd_delta(ytd, stub)

    # Mark voided
    prior_status = run.status
    run.status = "voided"
    run.voided_at = datetime.now(timezone.utc)
    run.void_reason = body.reason

    # Audit
    audit = PayrollAuditLog(
        owner_id=current_user.id,
        entity_type="pay_run",
        entity_id=run.id,
        action="voided",
        actor_user_id=current_user.id,
        actor_role="admin",
        before_state={"status": prior_status},
        after_state={"status": "voided", "reason": body.reason},
        notes=f"Voided pay run (prior status: {prior_status})",
    )
    db.add(audit)

    await db.commit()
    await db.refresh(run)
    return _run_to_response(run)


# =============================================================
# T4 preview formatters (SIN, province code, postal code)
# =============================================================
_T4_PROVINCE_MAP = {
    "alberta": "AB", "british columbia": "BC", "manitoba": "MB",
    "new brunswick": "NB", "newfoundland and labrador": "NL",
    "newfoundland": "NL", "labrador": "NL",
    "nova scotia": "NS", "northwest territories": "NT",
    "nunavut": "NU", "ontario": "ON",
    "prince edward island": "PE", "quebec": "QC", "qu\u00e9bec": "QC",
    "saskatchewan": "SK", "yukon": "YT",
}

def _t4_fmt_province(v):
    if not v:
        return ""
    v = v.strip()
    if len(v) == 2:
        return v.upper()
    return _T4_PROVINCE_MAP.get(v.lower(), v)

def _t4_fmt_sin(v):
    if not v:
        return ""
    digits = "".join(c for c in str(v) if c.isdigit())
    if len(digits) == 9:
        return f"{digits[0:3]}-{digits[3:6]}-{digits[6:9]}"
    return v.strip()

def _t4_fmt_postal(v):
    if not v:
        return ""
    v = v.strip().upper().replace(" ", "")
    if len(v) == 6:
        return f"{v[0:3]} {v[3:6]}"
    return v


@router.get("/taxes/t4-preview")
async def get_t4_preview(
    year: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    T4 employer slips preview data for the given tax year.

    Aggregates all finalized, non-voided paycheques per employee where
    pay_date falls within the tax year. Returns employer info plus one
    row per employee containing CRA T4 box values (14, 16, 16A, 18, 22,
    24, 26, 45, etc.).

    Response shape matches the T4EmployerSlips React component's props:
      { year, employer: {...}, employees: [{...}, ...] }

    Money box values are floats or null. Null renders as an empty box on
    the printed slip. Only employees with at least one non-voided pay
    stub in the tax year are included.
    """
    from datetime import date
    from app.models.models import CompanyProfile

    if year < 2020 or year > 2030:
        raise HTTPException(status_code=400, detail="year must be between 2020 and 2030")

    period_start = date(year, 1, 1)
    period_end = date(year, 12, 31)

    # ============================================================
    # 1) Company profile (employer info + Box 54 CRA account)
    # ============================================================
    comp_res = await db.execute(
        select(CompanyProfile).where(CompanyProfile.user_id == current_user.id)
    )
    company = comp_res.scalar_one_or_none()

    if not company:
        raise HTTPException(status_code=404, detail="Company profile not set")

    # Build CRA account: business_number + payroll_rp_account
    # e.g. "746043769" + "RP0001" = "746043769RP0001"
    bn = (company.business_number or "").strip()
    rp = (company.payroll_rp_account or "").strip()
    cra_account = bn + rp if bn and rp else (bn or rp or "")

    employer = {
        "name": company.company_name or "",
        "addr1": (company.address_street or "").strip(),
        "addr2": (company.address_city or "").strip(),
        "prov": _t4_fmt_province(company.province_state or ""),
        "postal": _t4_fmt_postal(company.address_postal_code or ""),
        "account": cra_account,
    }

    # ============================================================
    # 2) Aggregate paycheques per employee for the year
    # ============================================================
    # Group sums by employee_id
    result = await db.execute(
        select(
            PayStub.employee_id.label("employee_id"),
            func.coalesce(func.sum(PayStub.gross_pay), 0).label("gross"),
            func.coalesce(func.sum(PayStub.federal_tax), 0).label("federal_tax"),
            func.coalesce(func.sum(PayStub.provincial_or_state_tax), 0).label("provincial_tax"),
            func.coalesce(func.sum(PayStub.social_security_employee), 0).label("cpp_employee"),
            func.coalesce(func.sum(PayStub.social_security_2_employee), 0).label("cpp2_employee"),
            func.coalesce(func.sum(PayStub.unemployment_employee), 0).label("ei_employee"),
        )
        .select_from(PayStub)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(
            PayRun.owner_id == current_user.id,
            PayRun.pay_date >= period_start,
            PayRun.pay_date <= period_end,
            PayStub.voided == False,
            PayRun.status != "voided",
        )
        .group_by(PayStub.employee_id)
    )
    stub_agg = {row.employee_id: row for row in result.all()}

    if not stub_agg:
        return {"year": year, "employer": employer, "employees": []}

    # ============================================================
    # 3) Fetch employee records for those with paycheques
    # ============================================================
    employee_ids = list(stub_agg.keys())
    emp_res = await db.execute(
        select(Employee).where(Employee.id.in_(employee_ids))
    )
    employees = emp_res.scalars().all()

    # ============================================================
    # 4) Build the per-employee slip data
    # ============================================================
    def r2(v):
        """Round to 2 decimals; return None for zero (blank on the slip)."""
        v = float(v or 0)
        return round(v, 2) if v > 0 else None

    result_employees = []
    for emp in employees:
        agg = stub_agg.get(emp.id)
        if not agg:
            continue

        gross = float(agg.gross or 0)
        federal_tax = float(agg.federal_tax or 0)
        provincial_tax = float(agg.provincial_tax or 0)
        income_tax_deducted = federal_tax + provincial_tax

        result_employees.append({
            "id": str(emp.id),
            "last": (emp.last_name or "").strip(),
            "first": (emp.first_name or "").strip(),
            "init": "",  # middle initial - not tracked separately
            "sin": _t4_fmt_sin(emp.sin_or_ssn or ""),
            "addr1": (emp.address_line1 or "").strip(),
            "addr2": (emp.city or "").strip(),
            "province": _t4_fmt_province(emp.province_or_state or ""),
            "postal": _t4_fmt_postal(emp.postal_or_zip or ""),
            # Money boxes (None renders as blank on the slip)
            "b14": r2(gross),                        # Employment income
            "b16": r2(agg.cpp_employee),             # CPP contributions
            "b16A": r2(agg.cpp2_employee),           # Second CPP (CPP2)
            "b17": None,                             # QPP (Quebec only)
            "b17A": None,                            # Second QPP (Quebec only)
            "b18": r2(agg.ei_employee),              # EI premiums
            "b20": None,                             # RPP contributions - not tracked
            "b22": r2(income_tax_deducted),          # Income tax deducted
            "b24": r2(gross),                        # EI insurable earnings (proxy)
            "b26": r2(gross),                        # CPP pensionable earnings (proxy)
            "b44": None,                             # Union dues - not tracked
            "b46": None,                             # Charitable donations - not tracked
            "b50": None,                             # RPP registration - not tracked
            "b52": None,                             # Pension adjustment - not tracked
            "b55": None,                             # PPIP premiums (Quebec only)
            "b56": None,                             # PPIP insurable (Quebec only)
            "b45": (emp.dental_benefit_code or "1"), # Dental benefits code
        })

    # Sort alphabetically by last name for consistent order
    result_employees.sort(key=lambda e: (e["last"].upper(), e["first"].upper()))

    return {
        "year": year,
        "employer": employer,
        "employees": result_employees,
    }


@router.get("/taxes/t4-sum-preview")
async def get_t4_sum_preview(
    year: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    T4 Summary (T4 SUM 25) preview data for the given tax year.

    Aggregates YTD paycheque totals across ALL employees plus
    remittances from archived PD7A forms, and computes the CRA
    T4 Summary box values (14, 16, 16A, 27, 27A, 18, 19, 22,
    80, 82, difference, 84, 86, 88).

    Reconciliation performed server-side to guarantee:
      Box 80 = 16 + 16A + 27 + 27A + 18 + 19 + 22
      Difference = Box 80 - Box 82
      Box 84 = -Difference when negative (overpayment)
      Box 86 = Difference when positive (balance due)

    Response shape matches the T4Summary React component's props.
    """
    from datetime import date
    from app.models.models import CompanyProfile, ArchivedForm

    if year < 2020 or year > 2030:
        raise HTTPException(status_code=400, detail="year must be between 2020 and 2030")

    period_start = date(year, 1, 1)
    period_end = date(year, 12, 31)

    # ============================================================
    # 1) Company profile
    # ============================================================
    comp_res = await db.execute(
        select(CompanyProfile).where(CompanyProfile.user_id == current_user.id)
    )
    company = comp_res.scalar_one_or_none()

    if not company:
        raise HTTPException(status_code=404, detail="Company profile not set")

    bn = (company.business_number or "").strip()
    rp = (company.payroll_rp_account or "").strip()
    cra_account = bn + rp if bn and rp else (bn or rp or "")

    employer = {
        "name": company.company_name or "",
        "addr1": (company.address_street or "").strip(),
        "addr2": (company.address_city or "").strip(),
        "prov": _t4_fmt_province(company.province_state or ""),
        "postal": _t4_fmt_postal(company.address_postal_code or ""),
        "account": cra_account,
    }

    # ============================================================
    # 2) Aggregate paycheque totals across ALL employees for the year
    # ============================================================
    stubs_res = await db.execute(
        select(
            func.count(func.distinct(PayStub.employee_id)).label("slip_count"),
            func.coalesce(func.sum(PayStub.gross_pay), 0).label("gross"),
            func.coalesce(func.sum(PayStub.social_security_employee), 0).label("cpp_employee"),
            func.coalesce(func.sum(PayStub.social_security_2_employee), 0).label("cpp2_employee"),
            func.coalesce(func.sum(PayStub.social_security_employer), 0).label("cpp_employer"),
            func.coalesce(func.sum(PayStub.unemployment_employee), 0).label("ei_employee"),
            func.coalesce(func.sum(PayStub.unemployment_employer), 0).label("ei_employer"),
            func.coalesce(func.sum(PayStub.federal_tax), 0).label("federal_tax"),
            func.coalesce(func.sum(PayStub.provincial_or_state_tax), 0).label("provincial_tax"),
        )
        .select_from(PayStub)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(
            PayRun.owner_id == current_user.id,
            PayRun.pay_date >= period_start,
            PayRun.pay_date <= period_end,
            PayStub.voided == False,
            PayRun.status != "voided",
        )
    )
    totals = stubs_res.first()

    slip_count = int(totals.slip_count or 0)

    def _round2(v):
        return round(float(v or 0), 2)

    b14 = _round2(totals.gross)                # Employment income
    b16 = _round2(totals.cpp_employee)         # Employees' CPP
    b16A = _round2(totals.cpp2_employee)       # Employees' second CPP
    b27 = _round2(totals.cpp_employer)         # Employer's CPP
    b27A = 0.0                                 # Employer's second CPP (not tracked separately)
    b18 = _round2(totals.ei_employee)          # Employees' EI premiums
    b19 = _round2(totals.ei_employer)          # Employer's EI premiums
    b22 = _round2(float(totals.federal_tax or 0) + float(totals.provincial_tax or 0))

    # Box 80: total deductions reported = 16 + 16A + 27 + 27A + 18 + 19 + 22
    b80 = round(b16 + b16A + b27 + b27A + b18 + b19 + b22, 2)

    # ============================================================
    # 3) Box 82: remittances - sum of archived PD7A current_payment
    #    for the year
    # ============================================================
    remitt_res = await db.execute(
        select(ArchivedForm).where(
            ArchivedForm.user_id == current_user.id,
            ArchivedForm.country == "CA",
            ArchivedForm.form_type == "PD7A",
            ArchivedForm.period_start >= period_start,
            ArchivedForm.period_end <= period_end,
        )
    )
    archived = remitt_res.scalars().all()

    b82 = 0.0
    for form in archived:
        data = form.form_data or {}
        payment = data.get("current_payment") or 0
        try:
            b82 += float(payment)
        except (TypeError, ValueError):
            pass
    b82 = round(b82, 2)

    # ============================================================
    # 4) Difference, Box 84 (overpayment) or Box 86 (balance due)
    # ============================================================
    difference = round(b80 - b82, 2)
    if difference > 0:
        b84 = None
        b86 = difference
    elif difference < 0:
        b84 = abs(difference)
        b86 = None
    else:
        b84 = None
        b86 = None

    # Helper to convert 0 to None (blank cell on the form)
    def _blank_if_zero(v):
        return v if (v is not None and v > 0) else None

    summary = {
        "slips": slip_count,
        "b14": _blank_if_zero(b14),
        "b16": _blank_if_zero(b16),
        "b16A": _blank_if_zero(b16A),
        "b27": _blank_if_zero(b27),
        "b27A": _blank_if_zero(b27A),
        "b18": _blank_if_zero(b18),
        "b19": _blank_if_zero(b19),
        "b22": _blank_if_zero(b22),
        "b20": None,   # RPP contributions - not tracked
        "b52": None,   # Pension adjustment - not tracked
        "b80": _blank_if_zero(b80),
        "b82": _blank_if_zero(b82),
        "difference": abs(difference) if difference != 0 else None,
        "b84": b84,
        "b86": b86,
    }

    # ============================================================
    # 5) Contact: current user
    # ============================================================
    contact_name = ""
    if hasattr(current_user, "full_name") and current_user.full_name:
        contact_name = current_user.full_name
    elif hasattr(current_user, "first_name") or hasattr(current_user, "last_name"):
        first = getattr(current_user, "first_name", "") or ""
        last = getattr(current_user, "last_name", "") or ""
        contact_name = (first + " " + last).strip()
    elif getattr(current_user, "email", None):
        contact_name = current_user.email.split("@")[0]

    # Parse phone from company profile
    phone_raw = (company.phone or "").strip()
    area_code = ""
    phone_number = ""
    digits_only = "".join(c for c in phone_raw if c.isdigit())
    if len(digits_only) >= 10:
        area_code = digits_only[0:3]
        phone_number = digits_only[3:6] + "-" + digits_only[6:10]
    elif len(digits_only) == 7:
        phone_number = digits_only[0:3] + "-" + digits_only[3:7]
    else:
        phone_number = phone_raw

    contact = {
        "name": contact_name,
        "areaCode": area_code,
        "phone": phone_number,
        "ext": "",
    }

    return {
        "year": year,
        "employer": employer,
        "summary": summary,
        "contact": contact,
    }# ============================================================
# T4 employer slips PDF endpoint
# Returns a printable PDF version of the T4 slips, with real
# BrightCare data, using WeasyPrint to render HTML+CSS to PDF.
# ============================================================

T4_EMPLOYER_SLIPS_HTML_TEMPLATE = """<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  @page { size: letter; margin: 0.35in; }
  body { font-family: Arial, Helvetica, sans-serif; color: #000; font-size: 8.5pt; margin: 0; padding: 0; }

  .page { page-break-after: always; }
  .page:last-child { page-break-after: auto; }

  /* Each slip is a bordered box on the page */
  .slip {
    border: 1px solid #000;
    position: relative;
    padding: 6px 8px 7px 22px;
    margin-bottom: 22px;
  }
  .cutline {
    border-top: 1px dashed #666;
    margin: 18px 0;
  }

  /* Vertical Protected B strip on left of each slip */
  .vprot {
    position: absolute;
    left: 0; top: 0; bottom: 0; width: 18px;
    border-right: 1px solid #000;
    text-align: center;
    font-size: 6pt;
    padding: 6px 2px;
    writing-mode: vertical-rl;
    transform: rotate(180deg);
  }
  .vprot .prot-a { display: block; margin-bottom: 20px; font-weight: bold; }
  .vprot .prot-b { display: block; }

  /* HEADER (top band): 3 cells - employer / CRA / T4 title */
  .topband {
    display: table;
    width: 100%;
    border-collapse: collapse;
    table-layout: fixed;
  }
  .topband > div {
    display: table-cell;
    vertical-align: top;
  }
  .cell-employer {
    border: 1px solid #000;
    padding: 4px 6px;
    width: 40%;
    height: 62px;
  }
  .cell-cra {
    border-bottom: 1px solid #000;
    border-right: 1px solid #000;
    width: 33%;
    padding: 3px 5px;
  }
  .cell-t4title {
    text-align: right;
    padding: 2px 4px 3px;
    border-bottom: 1px solid #000;
  }

  .cap-label {
    font-size: 6.5pt;
    line-height: 1.15;
    margin-bottom: 3px;
  }
  .emp-addr {
    font-family: "Courier New", monospace;
    font-size: 10pt;
    line-height: 1.5;
  }

  /* CRA cell content */
  .cra-row1 {
    display: table;
    width: 100%;
    margin-bottom: 4px;
  }
  .cra-flag-cell {
    display: table-cell;
    width: 38px;
    vertical-align: middle;
  }
  .cra-name-cell {
    display: table-cell;
    vertical-align: middle;
    font-size: 7pt;
    line-height: 1.2;
    padding-left: 5px;
  }
  .flag {
    display: inline-block;
    width: 32px;
    height: 17px;
    background: linear-gradient(to right, #000 25%, #fff 25%, #fff 75%, #000 75%);
    border: 0.5px solid #999;
    text-align: center;
    line-height: 17px;
    font-size: 13px;
    color: #000;
  }
  .cra-row2 {
    display: table;
    width: 100%;
  }
  .yrlbl {
    display: table-cell;
    font-size: 6.5pt;
    line-height: 1.15;
    vertical-align: middle;
    padding-right: 8px;
    width: 55%;
  }
  .yrbox {
    display: table-cell;
    vertical-align: middle;
  }
  .yrbox span {
    display: inline-block;
    border: 1px solid #000;
    padding: 2px 10px;
    font-size: 12pt;
    font-family: "Courier New", monospace;
  }

  /* T4 title cell */
  .t4-big { font-size: 22pt; font-weight: bold; line-height: 1; }
  .t4-sub { font-size: 8.5pt; font-weight: bold; margin-top: 2px; }
  .t4-sub-fr { font-size: 8pt; font-style: italic; }

  /* BODY: 3-column layout below the top band */
  .bodyrow {
    display: table;
    width: 100%;
    table-layout: fixed;
    margin-top: 0;
  }
  .leftcol {
    display: table-cell;
    width: 40%;
    vertical-align: top;
  }
  .midcol {
    display: table-cell;
    width: 15%;
    vertical-align: top;
  }
  .rightcol {
    display: table-cell;
    width: 45%;
    vertical-align: top;
  }

  /* Box 54 (Employer account number) */
  .box54 {
    border: 1px solid #000;
    border-top: none;
    display: table;
    width: 100%;
    min-height: 28px;
  }
  .box54-num {
    display: table-cell;
    border-right: 1px solid #000;
    font-weight: bold;
    font-size: 9pt;
    padding: 2px 5px;
    vertical-align: top;
    width: 22px;
  }
  .box54-body {
    display: table-cell;
    padding: 2px 6px;
  }
  .box54-lbl { font-size: 6.5pt; }
  .box54-val {
    font-family: "Courier New", monospace;
    font-size: 12pt;
    letter-spacing: 1px;
    margin-top: 2px;
  }

  /* SIN + Exempt row */
  .sin-exempt-row {
    display: table;
    width: 100%;
    table-layout: fixed;
  }
  .sin-cell, .exempt-cell {
    display: table-cell;
    vertical-align: top;
  }
  .sin-cell { width: 55%; }
  .exempt-cell { width: 45%; }

  .field-lbl {
    text-align: center;
    font-size: 6pt;
    line-height: 1.1;
    padding: 2px 2px 0;
    min-height: 14px;
  }
  .field-box {
    border: 1px solid #000;
    border-top: none;
    display: table;
    width: 100%;
    min-height: 20px;
  }
  .field-num {
    display: table-cell;
    border-right: 1px solid #000;
    font-weight: bold;
    font-size: 8pt;
    padding: 1px 3px;
    vertical-align: top;
    width: 18px;
  }
  .field-val {
    display: table-cell;
    text-align: left;
    font-family: "Courier New", monospace;
    font-size: 10pt;
    padding: 2px 4px;
    vertical-align: middle;
  }

  /* Exempt checkboxes */
  .exempt-box {
    border: 1px solid #000;
    border-top: none;
    display: table;
    width: 100%;
    min-height: 30px;
  }
  .exempt-num {
    display: table-cell;
    border-right: 1px solid #000;
    font-weight: bold;
    font-size: 9pt;
    padding: 2px 3px;
    vertical-align: top;
    width: 18px;
  }
  .exempt-cols {
    display: table-cell;
    text-align: center;
    font-size: 5.5pt;
    padding: 3px 0;
  }
  .exempt-cols > span {
    display: inline-block;
    margin: 0 4px;
    text-align: center;
  }
  .cbx {
    border: 1px solid #000;
    width: 12px;
    height: 10px;
    display: inline-block;
    margin: 1px auto;
    vertical-align: middle;
  }

  /* Employee name and address block */
  .empname {
    border: 1px solid #000;
    padding: 4px 6px;
    min-height: 82px;
    margin-top: 3px;
  }
  .empname-hdr {
    font-size: 6pt;
    display: table;
    width: 100%;
    margin: 3px 0;
  }
  .empname-hdr > span {
    display: table-cell;
    padding-right: 4px;
  }
  .nameline {
    border: 1px solid #000;
    display: table;
    width: 100%;
    padding: 3px 5px;
    font-family: "Courier New", monospace;
    font-size: 10pt;
    table-layout: fixed;
  }
  .nameline > span {
    display: table-cell;
  }
  .nameline .ln { width: 58%; }
  .nameline .fn { width: 34%; }
  .nameline .inch { width: 8%; text-align: right; }
  .empaddr {
    font-family: "Courier New", monospace;
    font-size: 10pt;
    margin-top: 8px;
    line-height: 1.6;
  }

  /* Middle column boxes (45, 10, 29) */
  .mid-box {
    display: block;
    margin-bottom: 0;
  }

  /* Right column money boxes - 8 rows of 2 boxes each */
  .money-row {
    display: table;
    width: 100%;
    table-layout: fixed;
  }
  .money-row > .money-box {
    display: table-cell;
    width: 50%;
    vertical-align: top;
  }
  .money-box {
    display: block;
  }
  .money-lbl {
    text-align: center;
    font-size: 5.5pt;
    line-height: 1;
    padding: 1px 2px 0;
    min-height: 14px;
  }
  .money-field {
    border: 1px solid #000;
    border-top: none;
    display: table;
    width: 100%;
    min-height: 18px;
    table-layout: fixed;
  }
  .money-field > span {
    display: table-cell;
    vertical-align: middle;
  }
  .money-num {
    border-right: 1px solid #000;
    font-weight: bold;
    font-size: 7.5pt;
    padding: 1px 2px;
    width: 22px;
  }
  .money-dollars {
    text-align: right;
    font-family: "Courier New", monospace;
    font-size: 10pt;
    padding: 1px 3px;
  }
  .money-cents {
    border-left: 1px solid #000;
    text-align: right;
    font-family: "Courier New", monospace;
    font-size: 10pt;
    padding: 1px 3px;
    width: 22px;
  }

  /* Other information section */
  .other-section {
    border: 1px solid #000;
    margin-top: 6px;
    padding: 4px 6px;
  }
  .other-title {
    font-size: 6.5pt;
    margin-bottom: 3px;
  }
  .other-row {
    display: table;
    width: 100%;
    table-layout: fixed;
    margin-top: 2px;
  }
  .other-cell {
    display: table-cell;
    padding-right: 5px;
    vertical-align: top;
    width: 33.33%;
  }
  .other-cell:last-child { padding-right: 0; }
  .other-header {
    font-size: 6pt;
    display: table;
    width: 100%;
  }
  .other-header > span {
    display: table-cell;
  }
  .other-boxes {
    display: table;
    width: 100%;
    margin-top: 1px;
  }
  .other-boxes > span {
    display: table-cell;
  }
  .other-boxnum {
    border: 1px solid #000;
    width: 34px;
    height: 18px;
    background: #fff;
  }
  .other-boxval {
    border: 1px solid #000;
    border-left: none;
    height: 18px;
    background: #fff;
  }

  /* Slip footer */
  .slip-foot {
    display: table;
    width: 100%;
    margin-top: 4px;
    font-size: 7.5pt;
  }
  .slip-foot > div {
    display: table-cell;
  }
  .slip-foot > div:last-child {
    text-align: right;
    font-family: "Courier New", monospace;
  }

  /* Reverse page (box explanations) */
  .reverse-page {
    padding: 30px;
    font-size: 9pt;
    line-height: 1.4;
  }
  .reverse-title {
    font-size: 12pt;
    font-weight: bold;
    margin-bottom: 12px;
  }
  .reverse-cols {
    column-count: 2;
    column-gap: 24px;
  }
  .reverse-box {
    margin-bottom: 4px;
    break-inside: avoid;
  }
  .reverse-note {
    border: 1px solid #000;
    padding: 8px 10px;
    margin-top: 12px;
    font-size: 8pt;
  }
</style>
</head>
<body>
{% for pair in pages %}
<div class="page">
{% for e in pair %}
<div class="slip">
  <div class="vprot">
    <span class="prot-a">Protected B when completed / Protégé B une fois rempli</span>
    <span class="prot-b">T4 (25)</span>
  </div>

  <!-- TOP BAND: Employer / CRA / T4 title -->
  <div class="topband">
    <div class="cell-employer">
      <div class="cap-label">Employer's name &ndash; Nom de l'employeur</div>
      <div class="emp-addr">
        {{ employer.name }}<br>{{ employer.addr1 }}<br>
        {{ employer.addr2 }} &nbsp;&nbsp;&nbsp; {{ employer.prov }} &nbsp; {{ employer.postal }}
      </div>
    </div>
    <div class="cell-cra">
      <div class="cra-row1">
        <div class="cra-flag-cell"><span class="flag">🍁</span></div>
        <div class="cra-name-cell">Canada Revenue Agency<br>Agence du revenu du Canada</div>
      </div>
      <div class="cra-row2">
        <span class="yrlbl">Year<br>Année</span>
        <span class="yrbox"><span>{{ year }}</span></span>
      </div>
    </div>
    <div class="cell-t4title">
      <div class="t4-big">T4</div>
      <div class="t4-sub">Statement of Remuneration Paid</div>
      <div class="t4-sub-fr">État de la rémunération payée</div>
    </div>
  </div>

  <!-- BODY: Left 40% / Mid 15% / Right 45% -->
  <div class="bodyrow">
    <!-- LEFT COLUMN -->
    <div class="leftcol">
      <!-- Box 54 -->
      <div class="box54">
        <span class="box54-num">54</span>
        <div class="box54-body">
          <div class="box54-lbl">Employer's account number &ndash; Numéro de compte de l'employeur</div>
          <div class="box54-val">{{ employer.account }}</div>
        </div>
      </div>

      <!-- SIN + Exempt -->
      <div class="sin-exempt-row">
        <div class="sin-cell">
          <div class="field-lbl">Social insurance number<br>Numéro d'assurance sociale</div>
          <div class="field-box">
            <span class="field-num">12</span>
            <span class="field-val">{{ e.sin }}</span>
          </div>
        </div>
        <div class="exempt-cell">
          <div class="field-lbl"><b>Exempt &ndash; Exemption</b></div>
          <div class="exempt-box">
            <span class="exempt-num">28</span>
            <div class="exempt-cols">
              <span>CPP<br>QPP<br><span class="cbx"></span></span>
              <span>EI<br>AE<br><span class="cbx"></span></span>
              <span>PPIP<br>RPAP<br><span class="cbx"></span></span>
            </div>
          </div>
        </div>
      </div>

      <!-- Employee name and address -->
      <div class="empname">
        <div class="cap-label">Employee's name and address &ndash; Nom et adresse de l'employé</div>
        <div class="empname-hdr">
          <span>Last name (capitals) / Nom de famille</span>
          <span>First name / Prénom</span>
          <span>Init.</span>
        </div>
        <div class="nameline">
          <span class="ln">{{ e.last }}</span>
          <span class="fn">{{ e.first }}</span>
          <span class="inch">{{ e.init }}</span>
        </div>
        <div class="empaddr">
          {{ e.addr1 }}<br>{{ e.addr2 }} &nbsp;&nbsp;&nbsp; {{ e.province }} &nbsp; {{ e.postal }}
        </div>
      </div>
    </div>

    <!-- MIDDLE COLUMN -->
    <div class="midcol">
      <div class="mid-box">
        <div class="field-lbl">Dental benefits<br>Prestations dentaires</div>
        <div class="field-box">
          <span class="field-num">45</span>
          <span class="field-val">{{ e.b45 or "" }}</span>
        </div>
      </div>
      <div class="mid-box">
        <div class="field-lbl">Province of employment<br>Province d'emploi</div>
        <div class="field-box">
          <span class="field-num">10</span>
          <span class="field-val">{{ e.province }}</span>
        </div>
      </div>
      <div class="mid-box">
        <div class="field-lbl">Employment code<br>Code d'emploi</div>
        <div class="field-box">
          <span class="field-num">29</span>
          <span class="field-val"></span>
        </div>
      </div>
    </div>

    <!-- RIGHT COLUMN: 8 rows of 2 money boxes -->
    <div class="rightcol">
      {% set rows = [
        [("14", "Employment income – Revenus d'emploi", e.b14), ("22", "Income tax deducted – Impôt retenu", e.b22)],
        [("16", "CPP contributions – RPC", e.b16), ("17", "QPP contributions – RRQ", e.b17)],
        [("16A", "Second CPP – RPC2", e.b16A), ("17A", "Second QPP – RRQ2", e.b17A)],
        [("24", "EI insurable earnings – Gains AE", e.b24), ("26", "CPP/QPP pensionable – Gains RPC/RRQ", e.b26)],
        [("18", "EI premiums – Cotisations AE", e.b18), ("44", "Union dues – Cotisations syndicales", e.b44)],
        [("20", "RPP contributions – RPA", e.b20), ("46", "Charitable donations – Dons", e.b46)],
        [("52", "Pension adjustment – Facteur d'équivalence", e.b52), ("50", "RPP registration number", e.b50)],
        [("55", "PPIP premiums – Cotisations RPAP", e.b55), ("56", "PPIP insurable earnings – Gains RPAP", e.b56)]
      ] %}
      {% for row in rows %}
      <div class="money-row">
        {% for num, label, val in row %}
        <div class="money-box">
          <div class="money-lbl">{{ label }}</div>
          <div class="money-field">
            <span class="money-num">{{ num }}</span>
            {% if val is not none %}
              {% set dollars = ("%.2f" % val).split(".")[0] %}
              {% set cents = ("%.2f" % val).split(".")[1] %}
              <span class="money-dollars">{{ dollars }}</span>
              <span class="money-cents">{{ cents }}</span>
            {% else %}
              <span class="money-dollars"></span>
              <span class="money-cents"></span>
            {% endif %}
          </div>
        </div>
        {% endfor %}
      </div>
      {% endfor %}
    </div>
  </div>

  <!-- Other information section -->
  <div class="other-section">
    <div class="other-title">Other information (see over) &ndash; Autres renseignements (voir au verso)</div>
    <div class="other-row">
      <div class="other-cell">
        <div class="other-header"><span>Box &ndash; Case</span><span>Amount &ndash; Montant</span></div>
        <div class="other-boxes"><span class="other-boxnum"></span><span class="other-boxval"></span></div>
      </div>
      <div class="other-cell">
        <div class="other-header"><span>Box &ndash; Case</span><span>Amount &ndash; Montant</span></div>
        <div class="other-boxes"><span class="other-boxnum"></span><span class="other-boxval"></span></div>
      </div>
      <div class="other-cell">
        <div class="other-header"><span>Box &ndash; Case</span><span>Amount &ndash; Montant</span></div>
        <div class="other-boxes"><span class="other-boxnum"></span><span class="other-boxval"></span></div>
      </div>
    </div>
    <div class="other-row">
      <div class="other-cell">
        <div class="other-header"><span>Box &ndash; Case</span><span>Amount &ndash; Montant</span></div>
        <div class="other-boxes"><span class="other-boxnum"></span><span class="other-boxval"></span></div>
      </div>
      <div class="other-cell">
        <div class="other-header"><span>Box &ndash; Case</span><span>Amount &ndash; Montant</span></div>
        <div class="other-boxes"><span class="other-boxnum"></span><span class="other-boxval"></span></div>
      </div>
      <div class="other-cell">
        <div class="other-header"><span>Box &ndash; Case</span><span>Amount &ndash; Montant</span></div>
        <div class="other-boxes"><span class="other-boxnum"></span><span class="other-boxval"></span></div>
      </div>
    </div>
  </div>

  <div class="slip-foot">
    <div><b>T4 (25)</b></div>
    <div>REV &nbsp; OSP</div>
  </div>
</div>
{% if not loop.last %}<div class="cutline"></div>{% endif %}
{% endfor %}
</div>
{% endfor %}

<!-- REVERSE PAGE: Box explanations bilingual -->
<div class="reverse-page">
  <div class="reverse-title">Report these amounts on your tax return. / Déclarez ces montants dans votre déclaration.</div>
  <div class="reverse-cols">
    <div class="reverse-box"><b>14</b> Employment income – Enter on line 10100. Revenus d'emploi – ligne 10100.</div>
    <div class="reverse-box"><b>16</b> Employee's CPP contributions – see line 30800. Cotisations RPC.</div>
    <div class="reverse-box"><b>16A</b> Employee's second CPP contributions (CPP2). Deuxièmes cotisations au RPC.</div>
    <div class="reverse-box"><b>17</b> Employee's QPP contributions. Cotisations RRQ.</div>
    <div class="reverse-box"><b>17A</b> Employee's second QPP contributions (QPP2). Deuxièmes cotisations au RRQ.</div>
    <div class="reverse-box"><b>18</b> Employee's EI premiums – line 31200. Cotisations AE – ligne 31200.</div>
    <div class="reverse-box"><b>20</b> RPP contributions – line 20700. Cotisations RPA.</div>
    <div class="reverse-box"><b>22</b> Income tax deducted – line 43700. Impôt retenu – ligne 43700.</div>
    <div class="reverse-box"><b>24</b> EI insurable earnings. Gains assurables d'AE.</div>
    <div class="reverse-box"><b>26</b> CPP/QPP pensionable earnings. Gains ouvrant droit à pension.</div>
    <div class="reverse-box"><b>44</b> Union dues – line 21200. Cotisations syndicales – ligne 21200.</div>
    <div class="reverse-box"><b>45</b> Employer-offered dental benefits. Prestations dentaires.</div>
    <div class="reverse-box"><b>46</b> Charitable donations. Dons de bienfaisance.</div>
    <div class="reverse-box"><b>50</b> RPP or DPSP registration number. N° d'agrément RPA/RPDB.</div>
    <div class="reverse-box"><b>52</b> Pension adjustment – line 20600. Facteur d'équivalence – ligne 20600.</div>
    <div class="reverse-box"><b>54</b> Employer's account number. Numéro de compte de l'employeur.</div>
    <div class="reverse-box"><b>55</b> PPIP premiums. Cotisations RPAP.</div>
    <div class="reverse-box"><b>56</b> PPIP insurable earnings. Gains assurables RPAP.</div>
  </div>
  <div class="reverse-note">
    <b>Do not report these amounts on your tax return.</b> For CRA use only. (Amounts in boxes 30, 32, 34, 36, 38, 40, 57, 58, 59, 60, 86 and 90 are already included in box 14.)<br>
    <b>Ne déclarez pas ces montants.</b> À l'usage de l'ARC seulement.
  </div>
</div>
</body>
</html>
"""


@router.get("/taxes/t4-employer-slips.pdf")
async def get_t4_employer_slips_pdf(
    year: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Generate T4 employer slips as a downloadable PDF.

    Uses the same data aggregation as GET /taxes/t4-preview, renders it
    through a Jinja2 HTML template, and converts to PDF via WeasyPrint.

    Returns application/pdf inline so browsers display it in a new tab
    with their native PDF viewer (which has proper Print + Download).
    """
    from datetime import date
    from fastapi.responses import Response
    from jinja2 import Template
    from weasyprint import HTML
    from app.models.models import CompanyProfile

    if year < 2020 or year > 2030:
        raise HTTPException(status_code=400, detail="year must be between 2020 and 2030")

    period_start = date(year, 1, 1)
    period_end = date(year, 12, 31)

    # ==========================================================
    # 1) Company profile
    # ==========================================================
    comp_res = await db.execute(
        select(CompanyProfile).where(CompanyProfile.user_id == current_user.id)
    )
    company = comp_res.scalar_one_or_none()
    if not company:
        raise HTTPException(status_code=404, detail="Company profile not set")

    bn = (company.business_number or "").strip()
    rp = (company.payroll_rp_account or "").strip()
    cra_account = bn + rp if bn and rp else (bn or rp or "")

    employer = {
        "name": company.company_name or "",
        "addr1": (company.address_street or "").strip(),
        "addr2": (company.address_city or "").strip(),
        "prov": _t4_fmt_province(company.province_state or ""),
        "postal": _t4_fmt_postal(company.address_postal_code or ""),
        "account": cra_account,
    }

    # ==========================================================
    # 2) Aggregate paycheque data per employee
    # ==========================================================
    result = await db.execute(
        select(
            PayStub.employee_id.label("employee_id"),
            func.coalesce(func.sum(PayStub.gross_pay), 0).label("gross"),
            func.coalesce(func.sum(PayStub.federal_tax), 0).label("federal_tax"),
            func.coalesce(func.sum(PayStub.provincial_or_state_tax), 0).label("provincial_tax"),
            func.coalesce(func.sum(PayStub.social_security_employee), 0).label("cpp_employee"),
            func.coalesce(func.sum(PayStub.social_security_2_employee), 0).label("cpp2_employee"),
            func.coalesce(func.sum(PayStub.unemployment_employee), 0).label("ei_employee"),
        )
        .select_from(PayStub)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(
            PayRun.owner_id == current_user.id,
            PayRun.pay_date >= period_start,
            PayRun.pay_date <= period_end,
            PayStub.voided == False,
            PayRun.status != "voided",
        )
        .group_by(PayStub.employee_id)
    )
    stub_agg = {row.employee_id: row for row in result.all()}

    if not stub_agg:
        raise HTTPException(status_code=404, detail=f"No paycheques found for {year}")

    # ==========================================================
    # 3) Fetch employee records and build per-employee slip data
    # ==========================================================
    employee_ids = list(stub_agg.keys())
    emp_res = await db.execute(select(Employee).where(Employee.id.in_(employee_ids)))
    employees = emp_res.scalars().all()

    def r2(v):
        v = float(v or 0)
        return round(v, 2) if v > 0 else None

    employee_slips = []
    for emp in employees:
        agg = stub_agg.get(emp.id)
        if not agg:
            continue
        gross = float(agg.gross or 0)
        federal_tax = float(agg.federal_tax or 0)
        provincial_tax = float(agg.provincial_tax or 0)
        income_tax_deducted = federal_tax + provincial_tax

        employee_slips.append({
            "last": (emp.last_name or "").strip(),
            "first": (emp.first_name or "").strip(),
            "init": "",
            "sin": _t4_fmt_sin(emp.sin_or_ssn or ""),
            "addr1": (emp.address_line1 or "").strip(),
            "addr2": (emp.city or "").strip(),
            "province": _t4_fmt_province(emp.province_or_state or ""),
            "postal": _t4_fmt_postal(emp.postal_or_zip or ""),
            "b14": r2(gross),
            "b16": r2(agg.cpp_employee),
            "b16A": r2(agg.cpp2_employee),
            "b17": None,
            "b17A": None,
            "b18": r2(agg.ei_employee),
            "b20": None,
            "b22": r2(income_tax_deducted),
            "b24": r2(gross),
            "b26": r2(gross),
            "b44": None,
            "b46": None,
            "b50": None,
            "b52": None,
            "b55": None,
            "b56": None,
            "b45": (emp.dental_benefit_code or "1"),
        })

    employee_slips.sort(key=lambda e: (e["last"].upper(), e["first"].upper()))

    # Group into pairs (2 slips per page)
    pages = []
    for i in range(0, len(employee_slips), 2):
        pages.append(employee_slips[i:i+2])

    # ==========================================================
    # 4) Render HTML and generate PDF
    # ==========================================================
    template = Template(T4_EMPLOYER_SLIPS_HTML_TEMPLATE)
    html_content = template.render(
        year=year,
        employer=employer,
        pages=pages,
    )

    pdf_bytes = HTML(string=html_content).write_pdf()

    filename = f"T4-Employer-Slips-{year}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )# =============================================================
# T4 employer slips PDF - Playwright edition
# =============================================================
# Uses Playwright + Chromium headless to render the approved
# t4-slips-print.html template with real BrightCare data injected
# via window.__T4DATA__.
#
# This replaces the WeasyPrint version so the PDF matches the
# React preview exactly.
# =============================================================

import asyncio
import os
from pathlib import Path
from typing import Optional


# Module-level Playwright browser (lazy singleton)
_pw_browser = None
_pw_playwright = None
_pw_lock = asyncio.Lock()
_pw_semaphore = asyncio.Semaphore(int(os.getenv("T4_PDF_CONCURRENCY", "2")))


async def _get_browser():
    """Launch Chromium once and reuse it across requests."""
    global _pw_browser, _pw_playwright
    async with _pw_lock:
        if _pw_browser is None:
            from playwright.async_api import async_playwright
            _pw_playwright = await async_playwright().start()
            _pw_browser = await _pw_playwright.chromium.launch(
                headless=True,
                args=[
                    "--disable-dev-shm-usage",  # critical on small servers
                    "--no-sandbox",
                    "--disable-gpu",
                ],
            )
        return _pw_browser


async def shutdown_pw_browser():
    """Clean up Playwright on app shutdown."""
    global _pw_browser, _pw_playwright
    if _pw_browser is not None:
        await _pw_browser.close()
        _pw_browser = None
    if _pw_playwright is not None:
        await _pw_playwright.stop()
        _pw_playwright = None


def _money_pair(v):
    """Format a money value as ['dollars', 'cents'] pair, or None if blank."""
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if f == 0:
        return None
    parts = ("%.2f" % f).split(".")
    return [parts[0], parts[1]]


async def _render_t4_slips_pdf(data: dict) -> bytes:
    """
    Render the T4 slips PDF using Playwright.

    Loads the approved t4-slips-print.html template, injects data as
    window.__T4DATA__ (using add_init_script so it's set BEFORE the
    template's inline script runs), waits for slips to appear in the
    DOM, then generates PDF at Letter size.
    """
    # Path to the HTML template
    template_path = Path(__file__).resolve().parent.parent.parent / "templates" / "t4-slips-print.html"
    if not template_path.exists():
        raise HTTPException(
            status_code=500,
            detail=f"T4 template not found at {template_path}",
        )

    template_url = f"file://{template_path.absolute()}"

    async with _pw_semaphore:
        browser = await _get_browser()
        context = await browser.new_context()
        try:
            page = await context.new_page()

            # Inject data BEFORE the template's own script runs
            import json
            init_script = f"window.__T4DATA__ = {json.dumps(data)};"
            await page.add_init_script(init_script)

            # Load the template file
            await page.goto(template_url, wait_until="domcontentloaded")

            # Wait for at least one slip to appear (template builds the DOM)
            await page.wait_for_selector(".slip", timeout=5000)

            # Give the template's inline JS a moment to complete
            await page.wait_for_timeout(200)

            # Generate PDF
            pdf_bytes = await page.pdf(
                format="Letter",
                print_background=True,
                margin={"top": "0", "right": "0", "bottom": "0", "left": "0"},
                prefer_css_page_size=True,
            )
            return pdf_bytes
        finally:
            await context.close()


@router.get("/taxes/t4-employer-slips-v2.pdf")
async def get_t4_employer_slips_pdf_v2(
    year: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Generate T4 employer slips as a PDF using Playwright + Chromium
    to render the approved HTML template with real data.

    This produces a PDF that matches the React preview exactly since
    both use the same layout definitions.
    """
    from datetime import date
    from fastapi.responses import Response
    from app.models.models import CompanyProfile

    if year < 2020 or year > 2030:
        raise HTTPException(status_code=400, detail="year must be between 2020 and 2030")

    period_start = date(year, 1, 1)
    period_end = date(year, 12, 31)

    # ==========================================================
    # 1) Company profile
    # ==========================================================
    comp_res = await db.execute(
        select(CompanyProfile).where(CompanyProfile.user_id == current_user.id)
    )
    company = comp_res.scalar_one_or_none()
    if not company:
        raise HTTPException(status_code=404, detail="Company profile not set")

    bn = (company.business_number or "").strip()
    rp = (company.payroll_rp_account or "").strip()
    cra_account = bn + rp if bn and rp else (bn or rp or "")

    employer = {
        "name": company.company_name or "",
        "addr1": (company.address_street or "").strip(),
        "addr2": (company.address_city or "").strip(),
        "prov": _t4_fmt_province(company.province_state or ""),
        "postal": _t4_fmt_postal(company.address_postal_code or ""),
        "account": cra_account,
    }

    # ==========================================================
    # 2) Aggregate paycheque data per employee
    # ==========================================================
    result = await db.execute(
        select(
            PayStub.employee_id.label("employee_id"),
            func.coalesce(func.sum(PayStub.gross_pay), 0).label("gross"),
            func.coalesce(func.sum(PayStub.federal_tax), 0).label("federal_tax"),
            func.coalesce(func.sum(PayStub.provincial_or_state_tax), 0).label("provincial_tax"),
            func.coalesce(func.sum(PayStub.social_security_employee), 0).label("cpp_employee"),
            func.coalesce(func.sum(PayStub.social_security_2_employee), 0).label("cpp2_employee"),
            func.coalesce(func.sum(PayStub.unemployment_employee), 0).label("ei_employee"),
        )
        .select_from(PayStub)
        .join(PayRun, PayStub.pay_run_id == PayRun.id)
        .where(
            PayRun.owner_id == current_user.id,
            PayRun.pay_date >= period_start,
            PayRun.pay_date <= period_end,
            PayStub.voided == False,
            PayRun.status != "voided",
        )
        .group_by(PayStub.employee_id)
    )
    stub_agg = {row.employee_id: row for row in result.all()}

    if not stub_agg:
        raise HTTPException(status_code=404, detail=f"No paycheques found for {year}")

    # ==========================================================
    # 3) Fetch employee records
    # ==========================================================
    employee_ids = list(stub_agg.keys())
    emp_res = await db.execute(select(Employee).where(Employee.id.in_(employee_ids)))
    employees = emp_res.scalars().all()

    def r2(v):
        v = float(v or 0)
        return round(v, 2) if v > 0 else None

    # ==========================================================
    # 4) Build the data structure the template expects
    # ==========================================================
    employee_slips = []
    for emp in employees:
        agg = stub_agg.get(emp.id)
        if not agg:
            continue
        gross = float(agg.gross or 0)
        federal_tax = float(agg.federal_tax or 0)
        provincial_tax = float(agg.provincial_tax or 0)
        income_tax_deducted = federal_tax + provincial_tax

        employee_slips.append({
            "last": (emp.last_name or "").strip(),
            "first": (emp.first_name or "").strip(),
            "init": "",
            "sin": _t4_fmt_sin(emp.sin_or_ssn or ""),
            "addr1": (emp.address_line1 or "").strip(),
            "addr2": (emp.city or "").strip(),
            "province": _t4_fmt_province(emp.province_or_state or ""),
            "postal": _t4_fmt_postal(emp.postal_or_zip or ""),
            "b14": _money_pair(r2(gross)),
            "b16": _money_pair(r2(agg.cpp_employee)),
            "b16A": _money_pair(r2(agg.cpp2_employee)),
            "b17": None,
            "b17A": None,
            "b18": _money_pair(r2(agg.ei_employee)),
            "b20": None,
            "b22": _money_pair(r2(income_tax_deducted)),
            "b24": _money_pair(r2(gross)),
            "b26": _money_pair(r2(gross)),
            "b44": None,
            "b46": None,
            "b50": None,
            "b52": None,
            "b55": None,
            "b56": None,
            "b45": (emp.dental_benefit_code or "1"),
        })

    employee_slips.sort(key=lambda e: (e["last"].upper(), e["first"].upper()))

    data = {
        "year": year,
        "acct": cra_account,
        "employer": employer,
        "employees": employee_slips,
    }

    # ==========================================================
    # 5) Render PDF via Playwright
    # ==========================================================
    pdf_bytes = await _render_t4_slips_pdf(data)

    filename = f"T4-Employer-Slips-{year}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )